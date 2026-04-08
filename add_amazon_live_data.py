import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 加载V4文件 ──
wb = load_workbook(r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v4_20260327.xlsx")

# ── 样式函数 ──
def hstyle(cell, bg="1F3864", fg="FFFFFF", sz=11):
    cell.font = Font(name="微软雅黑", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dstyle(cell, bg="FFFFFF", bold=False, align="left", sz=10, color="000000"):
    cell.font = Font(name="微软雅黑", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, text, bg, fg="FFFFFF", sz=13, end_col="I", height=40):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=sz, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")

# ═══════════════════════════════════════════════════════════════════
# Sheet E — 亚马逊美国站实时数据（新增）
# ═══════════════════════════════════════════════════════════════════
ws_amz = wb.create_sheet("E-亚马逊美国站实时数据")
ws_amz.sheet_view.showGridLines = False
banner(ws_amz, 1, "亚马逊美国站（Amazon.com）实时销售数据 — Fanttik · Dreame · JISULIFE", "C00000", sz=13, end_col="K")

# 列宽
col_widths = [28, 12, 12, 16, 16, 14, 18, 16, 18, 18, 18]
for c, w in enumerate(col_widths, 1):
    ws_amz.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════════════════════════════════
# Fanttik 数据
# ═══════════════════════════════════════════════════════════════════
fanttik_data = [
    ["品牌", "产品型号", "评分", "评价数", "月销量", "价格(USD)", "主要卖点", "适用场景", "ASIN"],
    ["Fanttik", "X8 APEX 充气泵", "4.6", "6,403", "1000+", "$60", "1分钟快速充气·150PSI·LED双屏", "汽车/自行车/摩托车", "B09YD2D96V"],
    ["Fanttik", "X9 Pro 充气泵", "4.7", "2,785", "900+", "$60", "1分钟快充·150PSI·数字压力表", "小型汽车/电动自行车", "B0CJDZVDJN"],
    ["Fanttik", "X10 充气/鼓风机2合1", "4.7", "181", "300+", "$85", "2合1设计·充气+鼓风·移动电源", "汽车/气床垫/泳池浮舟", "B0DT11GNSB"],
    ["Fanttik", "X8 APEX EV版", "4.7", "1,011", "200+", "$60", "EV专用·充气速度快2倍", "电动汽车/汽车", "B0C36BBF7Y"],
    ["Fanttik", "X9 Ace 自行车泵", "4.4", "1,535", "1000+", "$40", "32s快充·150PSI·法式/Schrader阀", "山地自行车/公路自行车", "B0CLTTB86X"],
    ["Fanttik", "X10 Ace 迷你自行车泵", "4.5", "247", "700+", "$44", "120PSI·自动停止·迷你设计", "公路自行车", "B0F292DSXC"],
    ["Fanttik", "X10 Pro Max", "4.9", "36", "100+", "$90", "36s快充·SUV专用·强电池", "SUV/汽车/摩托车", "B0G29Q1NYC"],
    ["Fanttik", "X9 超便携", "4.7", "326", "50+", "$160", "3合1·气泵+电源站+手电筒", "皮卡车/RV/轻型卡车", "B0CQK4SRD8"],
]

# Fanttik标题
ws_amz.row_dimensions[3].height = 36
c = ws_amz.cell(row=3, column=1, value="▶ Fanttik 范泰克科技 — 亚马逊美国站数据（搜索结果：546条）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_amz.merge_cells("A3:K3")

for i, row in enumerate(fanttik_data, 4):
    ws_amz.row_dimensions[i].height = 26 if i == 4 else 24
    for c, val in enumerate(row, 1):
        cell = ws_amz.cell(row=i, column=c, value=val)
        if i == 4:
            hstyle(cell, bg="C00000", sz=10)
        else:
            bg = "FFE7E7" if i % 2 == 0 else "FFFFFF"
            dstyle(cell, bg=bg, bold=(c==1), align="center" if c > 1 else "left", sz=9)
        cell.border = tborder()

# ═══════════════════════════════════════════════════════════════════
# Dreame 数据
# ═══════════════════════════════════════════════════════════════════
dreame_data = [
    ["品牌", "产品型号", "评分", "评价数", "月销量", "价格(USD)", "主要卖点", "吸力(Pa)", "ASIN"],
    ["Dreame", "L40 Ultra Gen 2", "4.1", "151", "1000+", "$400", "25000Pa·可扩展侧刷·自清空底座", "25000", "B0FSJPFDPV"],
    ["Dreame", "C20 Plus", "4.6", "2,007", "50+", "$200", "90天自清空·6000Pa·LDS导航", "6000", "B0G918RT87"],
    ["Dreame", "L50 Ultra", "4.2", "397", "400+", "$800", "19500Pa·自清空·拖把自清洁", "19500", "B0F3J6FR1H"],
    ["Dreame", "D20 Plus", "4.3", "175", "300+", "$260", "13000Pa·5L防尘袋·宠物毛发", "13000", "B0F3C98LMB"],
    ["Dreame", "D10 Plus Gen 2", "4.2", "557", "700+", "$300", "90天清洁·6000Pa·LDS导航", "6000", "B0D6TTRTPL"],
    ["Dreame", "X50 Ultra Complete", "4.3", "601", "100+", "$900", "20000Pa·360°导航·边到边清洁", "20000", "B0F3HZFZBL"],
    ["Dreame", "L10s Pro Ultra", "4.2", "624", "600+", "$350", "7000Pa·120°F热水自清洁", "7000", "B0CVL2TT74"],
    ["Dreame", "X40 Ultra", "3.9", "769", "500+", "$540", "12000Pa·可升降拖把·自清空", "12000", "B0CXDXKSXP"],
    ["Dreame", "D30 Ultra", "4.6", "28", "50+", "$460", "25000Pa·100天自清空·边缘清洁", "25000", "B0G7C56BT5"],
]

dreame_start = 15
ws_amz.row_dimensions[dreame_start].height = 36
c = ws_amz.cell(row=dreame_start, column=1, value="▶ Dreame 追觅科技 — 亚马逊美国站数据（搜索结果：646条）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="7030A0")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_amz.merge_cells(f"A{dreame_start}:K{dreame_start}")

for i, row in enumerate(dreame_data, dreame_start + 1):
    ws_amz.row_dimensions[i].height = 26 if i == dreame_start + 1 else 24
    for c, val in enumerate(row, 1):
        cell = ws_amz.cell(row=i, column=c, value=val)
        if i == dreame_start + 1:
            hstyle(cell, bg="7030A0", sz=10)
        else:
            bg = "F3E5F5" if i % 2 == 0 else "FFFFFF"
            dstyle(cell, bg=bg, bold=(c==1), align="center" if c > 1 else "left", sz=9)
        cell.border = tborder()

# ═══════════════════════════════════════════════════════════════════
# JISULIFE 数据
# ═══════════════════════════════════════════════════════════════════
jisulife_data = [
    ["品牌", "产品型号", "评分", "评价数", "月销量", "价格(USD)", "主要卖点", "电池容量", "ASIN"],
    ["JISULIFE", "Life7 手持风扇", "4.7", "4,023", "3000+", "$24", "19.5小时续航·LED屏·5000mAh", "5000mAh", "B0CRDT715R"],
    ["JISULIFE", "Ultra2 手持风扇", "4.6", "530", "-", "$73", "25小时续航·9000mAh·5合1", "9000mAh", "B0F38H258L"],
    ["JISULIFE", "Life9 手持风扇", "4.6", "2,978", "1000+", "$29", "18小时续航·Turbo Power·5000mAh", "5000mAh", "B0CR3JJJTS"],
    ["JISULIFE", "3合1迷你风扇", "4.6", "76,130", "10000+", "$15", "最畅销·12-19小时·移动电源+手电筒", "4000mAh", "B07QK9C9KT"],
    ["JISULIFE", "Pro系列手持风扇", "4.6", "2,140", "1000+", "$60", "1-100速滚轮·金属机身·数字显示", "5000mAh", "B0C1SRTW9F"],
    ["JISULIFE", "涡轮风扇", "4.7", "9,422", "6000+", "$18", "16小时续航·4000mAh·5速", "4000mAh", "B09YCQTTTG"],
    ["JISULIFE", "台式风扇", "4.6", "2,424", "2000+", "$24", "180°可折叠·4500mAh·4速", "4500mAh", "B09PDC3JLB"],
    ["JISULIFE", "Life10s 手持风扇", "4.7", "559", "900+", "$24", "28小时续航·5000mAh·5速强风", "5000mAh", "B0F6LKL9JS"],
]

jisu_start = 28
ws_amz.row_dimensions[jisu_start].height = 36
c = ws_amz.cell(row=jisu_start, column=1, value="▶ JISULIFE 几素科技 — 亚马逊美国站数据（搜索结果：159条）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_amz.merge_cells(f"A{jisu_start}:K{jisu_start}")

for i, row in enumerate(jisulife_data, jisu_start + 1):
    ws_amz.row_dimensions[i].height = 26 if i == jisu_start + 1 else 24
    for c, val in enumerate(row, 1):
        cell = ws_amz.cell(row=i, column=c, value=val)
        if i == jisu_start + 1:
            hstyle(cell, bg="375623", sz=10)
        else:
            bg = "E2EFDA" if i % 2 == 0 else "FFFFFF"
            dstyle(cell, bg=bg, bold=(c==1), align="center" if c > 1 else "left", sz=9)
        cell.border = tborder()

# 说明
ws_amz.row_dimensions[40].height = 50
note = ws_amz.cell(row=40, column=1,
    value="数据来源：Amazon.com 美国站实时抓取（2026-03-27）| 价格为港币参考价，USD≈HKD/7.8 | 月销量为亚马逊显示的估算值 | ASIN为亚马逊产品唯一标识")
note.font = Font(name="微软雅黑", italic=True, size=9, color="595959")
note.fill = PatternFill("solid", fgColor="F2F2F2")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_amz.merge_cells("A40:K40")

ws_amz.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════════
out = r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v5_全球版_20260327.xlsx"
wb.save(out)
print(f"OK: {out}")
