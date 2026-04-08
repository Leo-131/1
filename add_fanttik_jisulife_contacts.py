# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 加载V6文件
wb = load_workbook(r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v6_全球全网版_20260327.xlsx")

# 样式函数
def hstyle(cell, bg="1F3864", fg="FFFFFF", sz=11):
    cell.font = Font(name="微软雅黑", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dstyle(cell, bg="FFFFFF", bold=False, align="left", sz=10, color="000000"):
    cell.font = Font(name="微软雅黑", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, text, bg, fg="FFFFFF", sz=13, end_col="K", height=40):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=sz, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")

# ═══════════════════════════════════════════════════════════════════
# Sheet G — Fanttik 全球经销商明细（新增）
# ═══════════════════════════════════════════════════════════════════
ws_f = wb.create_sheet("G-Fanttik全球经销商明细")
ws_f.sheet_view.showGridLines = False
banner(ws_f, 1, f"Fanttik 范泰克科技 — 全球经销商/代理商/KA联系方式明细 | 更新：{datetime.now().strftime('%Y-%m-%d')}", "C00000", sz=14, end_col="L")

# 列宽
for c, w in enumerate([14, 16, 14, 28, 22, 18, 16, 18, 14, 14, 14, 14], 1):
    ws_f.column_dimensions[get_column_letter(c)].width = w

# ── 一、品牌基本信息 ──
ws_f.row_dimensions[3].height = 36
c = ws_f.cell(row=3, column=1, value="一、Fanttik 品牌基本信息")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_f.merge_cells("A3:L3")

info_data = [
    ["公司名称", "深圳范泰克科技创新有限公司", "成立时间", "2020年"],
    ["品牌定位", "高端便携式汽车工具/户外装备/电动工具", "总部地址", "深圳市龙华区华侨城北站壹号大厦"],
    ["官网", "https://fanttik.com", "官网(德国)", "https://fanttik-innovation.de"],
    ["官网(英国)", "https://fanttik.uk", "客服邮箱", "support@fanttik.com"],
    ["营销邮箱", "marketing@fanttik.com", "批发咨询", "通过官网表单提交"],
    ["LinkedIn", "https://www.linkedin.com/company/fanttik/about", "Facebook", "https://www.facebook.com/fanttikofficial"],
    ["Instagram", "https://www.instagram.com/fanttik_official", "Twitter", "https://twitter.com/FanttikOfficial"],
    ["TikTok", "https://www.tiktok.com/@fanttik_official", "YouTube", "https://www.youtube.com/channel/UCCO_jSgj_5v0soeWII28n_g"],
]

for i, row in enumerate(info_data, 4):
    ws_f.row_dimensions[i].height = 26
    bg = "FFFFFF" if i % 2 == 0 else "FFE7E7"
    for c, val in enumerate(row, 1):
        cell = ws_f.cell(row=i, column=c, value=val)
        bold = (c % 2 == 1)
        dstyle(cell, bg=bg, bold=bold, sz=10)
        cell.border = tborder()

# ── 二、全球电商平台渠道 ──
ws_f.row_dimensions[14].height = 36
c = ws_f.cell(row=14, column=1, value="二、全球电商平台渠道（官方授权）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_f.merge_cells("A14:L14")

headers2 = ["国家/地区", "平台", "链接", "备注", "", "", "", "", "", "", "", ""]
for c, h in enumerate(headers2, 1):
    cell = ws_f.cell(row=15, column=c, value=h)
    hstyle(cell, bg="2E75B6", sz=10)

platform_data = [
    ["美国", "Amazon.com", "https://www.amazon.com/stores/FANTTIK/page/35DA621D-2547-47DE-AC21-55DD1E4040AE", "官方旗舰店"],
    ["美国", "Walmart", "https://www.walmart.com/seller/101066008", "官方授权店"],
    ["美国", "Costco", "https://www.costco.com/s?keyword=fanttik", "部分门店"],
    ["美国", "Target", "https://www.target.com/b/fanttik/-/N-q643le4oawf", "官方授权"],
    ["美国", "Home Depot", "https://www.homedepot.com/b/Fanttik/RV/N-5yc1vZvf5Z1z0mhhv", "官方授权"],
    ["美国", "Lowe's", "https://www.lowes.com/pd/Promounts/5013943253", "官方授权"],
    ["美国", "TikTok Shop", "https://vt.tiktok.com/ZT2AdCaHr/", "官方店铺"],
    ["德国", "Amazon.de", "https://www.amazon.de/-/en/stores/Fanttik/page/F4F6E28B-114B-4E85-85D9-04F07F8CD528", "官方旗舰店"],
    ["德国", "官网", "https://www.fanttik-innovation.de/", "德国独立站"],
    ["英国", "Amazon.co.uk", "https://www.amazon.co.uk/stores/Fanttik/page/2139187F-ACE5-4A2F-A8DA-0CF35EB6B533", "官方旗舰店"],
    ["英国", "官网", "https://fanttik.uk/", "英国独立站"],
    ["法国", "Amazon.fr", "https://www.amazon.fr/s?me=A22HDTVQZAOCR3", "官方店铺"],
    ["意大利", "Amazon.it", "https://www.amazon.it/s?me=A22HDTVQZAOCR3", "官方店铺"],
    ["西班牙", "Amazon.es", "https://www.amazon.es/s?me=A22HDTVQZAOCR3", "官方店铺"],
    ["加拿大", "Amazon.ca", "https://www.amazon.ca/stores/WelcometoFanttikCA/page/60FD366B-1922-4475-A254-3167EECE97BC", "官方店铺"],
    ["日本", "Amazon.co.jp", "https://www.amazon.co.jp/b?node=26286483051", "官方店铺"],
    ["澳大利亚", "Amazon.com.au", "https://www.amazon.com.au/s?me=A3NHVZ8N64UWP3", "官方店铺"],
]

for i, row in enumerate(platform_data, 16):
    ws_f.row_dimensions[i].height = 26
    bg = "FFFFFF" if i % 2 == 0 else "FFE7E7"
    for c, val in enumerate(row, 1):
        cell = ws_f.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=10)
        cell.border = tborder()
    ws_f.merge_cells(f"D{i}:L{i}")

# ── 三、批发/代理商合作 ──
ws_f.row_dimensions[36].height = 36
c = ws_f.cell(row=36, column=1, value="三、批发/代理商合作联系方式")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_f.merge_cells("A36:L36")

headers3 = ["合作类型", "联系方式", "邮箱", "申请流程", "回复时间", "优惠信息", "", "", "", "", "", ""]
for c, h in enumerate(headers3, 1):
    cell = ws_f.cell(row=37, column=c, value=h)
    hstyle(cell, bg="2E75B6", sz=10)

wholesale_data = [
    ["批量采购/企业礼品", "官网表单", "support@fanttik.com / marketing@fanttik.com", "官网填写表单，注明产品名称和数量", "24小时内", "量大从优，需询价"],
    ["分销商/代理商", "官网表单", "marketing@fanttik.com", "提交合作意向，等待审核", "24小时内", "授权经销商价格"],
    ["联盟营销(Affiliate)", "官网注册", "https://affiliates.fanttik.com/create-account", "在线注册，审核后开始推广", "3-5个工作日", "佣金比例需洽谈"],
    ["品牌大使(Ambassador)", "官网表单", "https://fanttik.com/pages/ambassador-signup", "提交申请，审核资格", "3-5个工作日", "专属折扣+佣金"],
]

for i, row in enumerate(wholesale_data, 38):
    ws_f.row_dimensions[i].height = 28
    bg = "FFFFFF" if i % 2 == 0 else "FFE7E7"
    for c, val in enumerate(row, 1):
        cell = ws_f.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=10)
        cell.border = tborder()
    ws_f.merge_cells(f"F{i}:L{i}")

# ── 四、KA大客户/企业采购 ──
ws_f.row_dimensions[45].height = 36
c = ws_f.cell(row=45, column=1, value="四、KA大客户/企业采购合作")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_f.merge_cells("A45:L45")

ka_data = [
    ["客户类型", "典型场景", "采购流程", "联系方式", "优势"],
    ["企业客户", "员工福利/团建礼品", "邮件询价→报价确认→签订合同→发货", "marketing@fanttik.com", "批量折扣/定制包装"],
    ["俱乐部/活动方", "赛事奖品/活动礼品", "提交活动信息→方案报价→采购执行", "marketing@fanttik.com", "赞助支持/品牌露出"],
    ["零售商/经销商", "线下门店销售", "提交资质审核→签订授权协议→首批采购", "marketing@fanttik.com", "区域保护/营销支持"],
    ["电商卖家", "亚马逊/eBay等平台销售", "提交店铺链接→审核授权→开始销售", "marketing@fanttik.com", "官方授权/价格保护"],
]

for i, row in enumerate(ka_data, 46):
    ws_f.row_dimensions[i].height = 28 if i == 46 else 26
    bg = "2E75B6" if i == 46 else ("FFFFFF" if i % 2 == 0 else "FFE7E7")
    for c, val in enumerate(row, 1):
        cell = ws_f.cell(row=i, column=c, value=val)
        if i == 46:
            hstyle(cell, bg="2E75B6", sz=10)
        else:
            dstyle(cell, bg=bg, bold=(c==1), sz=10)
        cell.border = tborder()

# ── 五、子品牌渠道 ──
ws_f.row_dimensions[53].height = 36
c = ws_f.cell(row=53, column=1, value="五、Fanttik子品牌渠道（FanttikRide / FanttikOutdoor）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_f.merge_cells("A53:L53")

sub_headers = ["子品牌", "产品线", "目标市场", "主要渠道", "官网"]
for c, h in enumerate(sub_headers, 1):
    cell = ws_f.cell(row=54, column=c, value=h)
    hstyle(cell, bg="2E75B6", sz=10)

sub_data = [
    ["FanttikRide", "儿童电动滑板车/骑行玩具", "美国/德国/英国/法国/西班牙/意大利/加拿大", "Amazon/Walmart/TikTok Shop", "fanttik.com/pages/fanttikride"],
    ["FanttikOutdoor", "帐篷/露营桌椅/户外装备", "美国", "Amazon/官网", "fanttik.com/pages/fanttikoutdoor"],
]

for i, row in enumerate(sub_data, 55):
    ws_f.row_dimensions[i].height = 28
    bg = "FFFFFF" if i % 2 == 0 else "FFE7E7"
    for c, val in enumerate(row, 1):
        cell = ws_f.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=10)
        cell.border = tborder()

ws_f.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════
# Sheet H — JISULIFE 全球经销商明细（新增）
# ═══════════════════════════════════════════════════════════════════
ws_j = wb.create_sheet("H-JISULIFE全球经销商明细")
ws_j.sheet_view.showGridLines = False
banner(ws_j, 1, f"JISULIFE 几素科技 — 全球经销商/代理商/KA联系方式明细 | 更新：{datetime.now().strftime('%Y-%m-%d')}", "375623", sz=14, end_col="L")

# 列宽
for c, w in enumerate([14, 16, 14, 28, 22, 18, 16, 18, 14, 14, 14, 14], 1):
    ws_j.column_dimensions[get_column_letter(c)].width = w

# ── 一、品牌基本信息 ──
ws_j.row_dimensions[3].height = 36
c = ws_j.cell(row=3, column=1, value="一、JISULIFE 几素品牌基本信息")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_j.merge_cells("A3:L3")

j_info_data = [
    ["公司名称", "深圳市几素科技有限公司", "成立时间", "2016年"],
    ["品牌定位", "手持风扇/挂脖风扇开创者·全球细分领域第一", "总部地址", "广东省深圳市"],
    ["官网", "https://jisulife.com", "市场覆盖", "40+国家"],
    ["客服邮箱", "support@jisulife.com", "客服电话", "+1 888-252-0666"],
    ["工作时间", "Monday to Friday: 9:00 am - 18:00 pm (PST)", "退款政策", "30天退款保证"],
    ["质保政策", "1年质保", "配送政策", "全球免邮"],
]

for i, row in enumerate(j_info_data, 4):
    ws_j.row_dimensions[i].height = 26
    bg = "FFFFFF" if i % 2 == 0 else "E2EFDA"
    for c, val in enumerate(row, 1):
        cell = ws_j.cell(row=i, column=c, value=val)
        bold = (c % 2 == 1)
        dstyle(cell, bg=bg, bold=bold, sz=10)
        cell.border = tborder()

# ── 二、业务合作联系方式（核心）──
ws_j.row_dimensions[12].height = 36
c = ws_j.cell(row=12, column=1, value="二、业务合作联系方式（按合作类型分类）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_j.merge_cells("A12:L12")

j_headers = ["合作类型", "邮箱", "适用场景", "申请流程", "回复时间", "备注", "", "", "", "", "", ""]
for c, h in enumerate(j_headers, 1):
    cell = ws_j.cell(row=13, column=c, value=h)
    hstyle(cell, bg="548235", sz=10)

j_contact_data = [
    ["⭐批发分销", "b2b@jisulife.com", "经销商/分销商/线下门店/电商平台卖家", "邮件联系，说明公司信息和采购需求", "24小时内", "首选渠道"],
    ["联盟营销", "affiliate@jisulife.com", "个人博主/内容创作者/网站站长", "邮件申请或官网注册", "3-5个工作日", "佣金模式"],
    ["网红合作", "influencer@jisulife.com", "Instagram/TikTok/YouTube网红", "邮件联系，附粉丝数据和媒体资料", "3-5个工作日", "产品置换/付费推广"],
    ["公关营销", "marketing@jisulife.com", "媒体采访/品牌合作/活动赞助", "邮件说明合作意向", "24小时内", "品牌层面合作"],
    ["客服支持", "support@jisulife.com", "产品咨询/售后问题/订单查询", "邮件或电话", "24小时内", "客服热线: +1 888-252-0666"],
]

for i, row in enumerate(j_contact_data, 14):
    ws_j.row_dimensions[i].height = 28
    bg = "FFFFFF" if i % 2 == 0 else "E2EFDA"
    color = "375623" if row[0].startswith("⭐") else "000000"
    for c, val in enumerate(row, 1):
        cell = ws_j.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=10, color=color if c==1 else "000000")
        cell.border = tborder()
    ws_j.merge_cells(f"F{i}:L{i}")

# ── 三、授权分销商计划详情 ──
ws_j.row_dimensions[22].height = 36
c = ws_j.cell(row=22, column=1, value="三、JISULIFE 授权分销商计划详情")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_j.merge_cells("A22:L22")

j_dist_headers = ["计划类型", "申请链接", "适用对象", "优势", "申请条件"]
for c, h in enumerate(j_dist_headers, 1):
    cell = ws_j.cell(row=23, column=c, value=h)
    hstyle(cell, bg="548235", sz=10)

j_dist_data = [
    ["线下经销商", "jisulife.com/pages/become-a-dealer", "实体店/零售商/批发商", "区域保护/营销支持/价格优势", "无需财务证明/信用审查，个人或企业均可"],
    ["线上联盟", "jisulife.com/pages/jisulife-affiliate-program", "电商卖家/网站站长/内容创作者", "零库存/佣金模式/官方素材支持", "有销售渠道或内容平台即可"],
]

for i, row in enumerate(j_dist_data, 24):
    ws_j.row_dimensions[i].height = 28
    bg = "FFFFFF" if i % 2 == 0 else "E2EFDA"
    for c, val in enumerate(row, 1):
        cell = ws_j.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=10)
        cell.border = tborder()

# ── 四、社交媒体 ──
ws_j.row_dimensions[29].height = 36
c = ws_j.cell(row=29, column=1, value="四、JISULIFE 社交媒体渠道")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_j.merge_cells("A29:L29")

j_social_headers = ["平台", "账号", "链接", "粉丝数(估)", "", "", "", "", "", "", "", ""]
for c, h in enumerate(j_social_headers, 1):
    cell = ws_j.cell(row=30, column=c, value=h)
    hstyle(cell, bg="548235", sz=10)

j_social_data = [
    ["TikTok", "@jisulife.official", "https://www.tiktok.com/@jisulife.official", "200万+"],
    ["Instagram", "@jisulife.official", "https://www.instagram.com/jisulife.official/", "50万+"],
    ["Facebook", "JisuLife.Official", "https://www.facebook.com/JisuLife.Official", "20万+"],
    ["YouTube", "JISULIFE Official", "https://www.youtube.com/channel/UC9gCraqBWS-D4rcQuHh0m6g", "10万+"],
    ["Twitter/X", "@JisulifeGlobal", "https://x.com/JisulifeGlobal", "3万+"],
    ["Pinterest", "JisuLifeGlobal", "https://www.pinterest.com/JisuLifeGlobal/", "N/A"],
]

for i, row in enumerate(j_social_data, 31):
    ws_j.row_dimensions[i].height = 26
    bg = "FFFFFF" if i % 2 == 0 else "E2EFDA"
    for c, val in enumerate(row, 1):
        cell = ws_j.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=10)
        cell.border = tborder()
    ws_j.merge_cells(f"D{i}:L{i}")

# ── 五、主要销售渠道 ──
ws_j.row_dimensions[40].height = 36
c = ws_j.cell(row=40, column=1, value="五、JISULIFE 主要销售渠道（全球）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_j.merge_cells("A40:L40")

j_channel_headers = ["渠道类型", "平台/渠道", "市场地位", "备注"]
for c, h in enumerate(j_channel_headers, 1):
    cell = ws_j.cell(row=41, column=c, value=h)
    hstyle(cell, bg="548235", sz=10)

j_channel_data = [
    ["电商", "Amazon.com", "⭐BSR #1 手持风扇类目", "官方旗舰店，月销10000+"],
    ["电商", "Amazon.de/fr/it/es", "BSR上榜", "欧洲多站点"],
    ["电商", "天猫旗舰店", "⭐销量第一 连续两年", "jisulife旗舰店"],
    ["电商", "京东自营", "官方授权", "jisulife京东自营旗舰店"],
    ["线下零售", "Walmart", "3000+门店", "美国本土"],
    ["线下零售", "Costco", "全球门店", "会员制卖场"],
    ["独立站", "jisulife.com", "DTC官网", "全球免邮"],
]

for i, row in enumerate(j_channel_data, 42):
    ws_j.row_dimensions[i].height = 26
    bg = "FFFFFF" if i % 2 == 0 else "E2EFDA"
    color = "375623" if "⭐" in row[2] else "000000"
    for c, val in enumerate(row, 1):
        cell = ws_j.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1 or "⭐" in str(val)), sz=10, color=color)
        cell.border = tborder()

ws_j.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════════
out = r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v7_全球联系方式明细_20260330.xlsx"
wb.save(out)
print(f"OK: {out}")
