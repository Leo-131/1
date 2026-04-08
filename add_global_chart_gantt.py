import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import datetime

# ── 加载现有文件 ──
wb = load_workbook(r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_20260325.xlsx")

# ── 样式函数 ──
def hstyle(cell, bg="1F3864", fg="FFFFFF", sz=11):
    cell.font = Font(name="微软雅黑", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dstyle(cell, bg="FFFFFF", bold=False, align="left", sz=10, color="000000"):
    cell.font = Font(name="微软雅黑", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, text, bg, fg="FFFFFF", sz=13, end_col="H", height=40):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=sz, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")

# ══════════════════════════════════════════════
# Sheet A — 全球经销商数据汇总（新增）
# ══════════════════════════════════════════════
ws_g = wb.create_sheet("全球经销商数据汇总")
ws_g.sheet_view.showGridLines = False
banner(ws_g, 1, "竞品品牌 全球经销商/分销商/零售商 数据汇总（2026年3月）", "1F3864", sz=14, end_col="I")

# 表头
h = ["品牌", "中国大陆", "中国港台", "日本/韩国", "东南亚", "欧洲", "北美", "澳洲/中东", "全球合计（估）"]
w = [18, 12, 12, 12, 12, 14, 14, 14, 14]
for c, (hh, ww) in enumerate(zip(h, w), 1):
    cell = ws_g.cell(row=2, column=c, value=hh)
    hstyle(cell, bg="2E75B6")
    ws_g.column_dimensions[get_column_letter(c)].width = ww
ws_g.row_dimensions[2].height = 38

# 数据（基于公开信息整理，含估算）
data = [
    # 品牌, 中国大陆, 港台, 日韩, 东南亚, 欧洲, 北美, 澳洲/中东, 合计
    ["Flextail 鱼尾",    "10+",  "2",  "3",  "5",  "8",  "12", "4",  "44+"],
    ["电小二/Jackery",   "50+",  "5",  "8",  "10", "15", "30", "8",  "126+"],
    ["正浩 EcoFlow",     "30+",  "4",  "6",  "8",  "20", "25", "10", "103+"],
    ["铂陆帝 BLUETTI",   "20+",  "3",  "5",  "6",  "18", "22", "8",  "82+"],
    ["安克 Anker",       "40+",  "5",  "10", "12", "20", "35", "10", "132+"],
    ["公牛 BULL",        "200+", "2",  "1",  "3",  "2",  "2",  "1",  "211+"],
    ["纽曼 Newsmy",      "80+",  "3",  "2",  "4",  "3",  "3",  "2",  "97+"],
    ["Go Outdoors",      "0",    "0",  "0",  "0",  "1",  "0",  "0",  "1（英国本土58店）"],
    ["Goal Zero",        "0",    "0",  "1",  "1",  "5",  "20", "3",  "30+"],
]

colors = ["FFFFFF", "EBF3FB"]
for i, row in enumerate(data, 3):
    bg = colors[i % 2]
    for c, val in enumerate(row, 1):
        cell = ws_g.cell(row=i, column=c, value=val)
        bold = (c == 1)
        color = "C00000" if val == "0" else ("1F497D" if c == 9 else "000000")
        dstyle(cell, bg=bg, bold=bold, align="center" if c > 1 else "left", sz=10, color=color)
        cell.border = tborder()
    ws_g.row_dimensions[i].height = 28

# 说明
ws_g.row_dimensions[13].height = 50
note = ws_g.cell(row=13, column=1,
    value="数据说明：以上数据基于公开电商平台（亚马逊/1688/速卖通/eBay）、品牌官网、行业报告及新闻整理，含估算。"
          "「经销商」包含官方授权代理商、分销商、零售商及电商平台授权店铺。"
          "Go Outdoors为英国本土零售商（58家实体店），非品牌商，不参与全球分销。")
note.font = Font(name="微软雅黑", italic=True, size=9, color="595959")
note.fill = PatternFill("solid", fgColor="F2F2F2")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_g.merge_cells("A13:I13")

ws_g.freeze_panes = "A3"

# ══════════════════════════════════════════════
# Sheet B — 数据可视化图表
# ══════════════════════════════════════════════
ws_chart = wb.create_sheet("经销商数量对比图表")
ws_chart.sheet_view.showGridLines = False
banner(ws_chart, 1, "各竞品品牌 全球经销商数量对比（数据可视化）", "1F3864", sz=14, end_col="N")

# 图表数据源（数值版，用于绘图）
chart_headers = ["品牌", "中国大陆", "日本/韩国", "东南亚", "欧洲", "北美", "澳洲/中东"]
chart_data = [
    ["Flextail 鱼尾",   10,  3,  5,  8,  12, 4],
    ["电小二/Jackery",  50,  8,  10, 15, 30, 8],
    ["正浩 EcoFlow",    30,  6,  8,  20, 25, 10],
    ["铂陆帝 BLUETTI",  20,  5,  6,  18, 22, 8],
    ["安克 Anker",      40,  10, 12, 20, 35, 10],
    ["公牛 BULL",       200, 1,  3,  2,  2,  1],
    ["纽曼 Newsmy",     80,  2,  4,  3,  3,  2],
    ["Goal Zero",       0,   1,  1,  5,  20, 3],
]

for c, h in enumerate(chart_headers, 1):
    cell = ws_chart.cell(row=3, column=c, value=h)
    hstyle(cell, bg="2E75B6")
    ws_chart.column_dimensions[get_column_letter(c)].width = 16
ws_chart.row_dimensions[3].height = 36

for i, row in enumerate(chart_data, 4):
    bg = "FFFFFF" if i % 2 == 0 else "EBF3FB"
    for c, val in enumerate(row, 1):
        cell = ws_chart.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), align="center" if c > 1 else "left", sz=10)
        cell.border = tborder()
    ws_chart.row_dimensions[i].height = 26

# ── 图表1：堆叠柱状图（全球分布）──
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "stacked"
chart1.overlap = 100
chart1.title = "各竞品品牌 全球经销商数量分布（堆叠）"
chart1.y_axis.title = "经销商数量（家）"
chart1.x_axis.title = "品牌"
chart1.style = 10
chart1.width = 22
chart1.height = 14

cats = Reference(ws_chart, min_col=1, min_row=4, max_row=11)
for col in range(2, 8):
    data_ref = Reference(ws_chart, min_col=col, min_row=3, max_row=11)
    chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats)
ws_chart.add_chart(chart1, "A14")

# ── 图表2：簇状柱状图（北美 vs 欧洲 vs 中国大陆）──
chart2 = BarChart()
chart2.type = "col"
chart2.grouping = "clustered"
chart2.title = "核心市场经销商数量对比（中国大陆 vs 欧洲 vs 北美）"
chart2.y_axis.title = "经销商数量（家）"
chart2.x_axis.title = "品牌"
chart2.style = 26
chart2.width = 22
chart2.height = 14

cats2 = Reference(ws_chart, min_col=1, min_row=4, max_row=11)
for col in [2, 5, 6]:  # 中国大陆、欧洲、北美
    data_ref2 = Reference(ws_chart, min_col=col, min_row=3, max_row=11)
    chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats2)
ws_chart.add_chart(chart2, "L14")

# ══════════════════════════════════════════════
# Sheet C — 项目进度甘特图
# ══════════════════════════════════════════════
ws_gantt = wb.create_sheet("市场调研甘特图")
ws_gantt.sheet_view.showGridLines = False
banner(ws_gantt, 1, "Flextail & Vollyc 竞品市场调研 — 项目进度甘特图（2026年4月-9月）", "1F3864", sz=14, end_col="T")

# 月份标题行
months = ["4月W1", "4月W2", "4月W3", "4月W4",
          "5月W1", "5月W2", "5月W3", "5月W4",
          "6月W1", "6月W2", "6月W3", "6月W4",
          "7月W1", "7月W2", "7月W3", "7月W4",
          "8月W1", "8月W2", "8月W3", "8月W4"]

# 列宽
ws_gantt.column_dimensions["A"].width = 28  # 任务名
ws_gantt.column_dimensions["B"].width = 12  # 负责人
ws_gantt.column_dimensions["C"].width = 10  # 优先级
for i, m in enumerate(months, 4):
    ws_gantt.column_dimensions[get_column_letter(i)].width = 5

# 表头行
ws_gantt.row_dimensions[2].height = 36
ws_gantt.cell(row=2, column=1, value="任务名称")
ws_gantt.cell(row=2, column=2, value="负责人")
ws_gantt.cell(row=2, column=3, value="优先级")
for c in [1, 2, 3]:
    hstyle(ws_gantt.cell(row=2, column=c), bg="1F3864")

for i, m in enumerate(months, 4):
    cell = ws_gantt.cell(row=2, column=i, value=m)
    bg = "2E75B6" if "W1" in m else ("375623" if "W2" in m else ("7030A0" if "W3" in m else "ED7D31"))
    hstyle(cell, bg=bg, sz=8)

# 月份合并标题
ws_gantt.row_dimensions[3].height = 22
month_labels = [("4月", 4, 7), ("5月", 8, 11), ("6月", 12, 15), ("7月", 16, 19), ("8月", 20, 23)]
for label, start, end in month_labels:
    ws_gantt.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
    cell = ws_gantt.cell(row=3, column=start, value=label)
    hstyle(cell, bg="1F3864", sz=10)

# 甘特图任务数据
# (任务名, 负责人, 优先级, 开始周(1-20), 持续周数, 颜色)
tasks = [
    # 阶段一：基础调研（4月）
    ("【阶段一】基础竞品调研",          "",     "",     1,  0,  "1F3864"),  # 分组标题
    ("1.1 Flextail全球渠道梳理",        "市场组", "🔴高", 1,  3,  "C00000"),
    ("1.2 Vollyc品牌核实与定位",        "市场组", "🔴高", 1,  2,  "C00000"),
    ("1.3 电小二/Jackery全球经销商调研", "市场组", "🟡中", 2,  3,  "ED7D31"),
    ("1.4 EcoFlow全球经销商调研",       "市场组", "🟡中", 2,  3,  "ED7D31"),
    ("1.5 BLUETTI全球经销商调研",       "市场组", "🟡中", 3,  3,  "ED7D31"),
    ("1.6 Go Outdoors破产进展跟踪",     "市场组", "🟢低", 3,  2,  "375623"),

    # 阶段二：深度调研（5月）
    ("【阶段二】深度市场调研",           "",     "",     5,  0,  "1F3864"),
    ("2.1 北美市场经销商实地调研",       "海外组", "🔴高", 5,  4,  "C00000"),
    ("2.2 欧洲市场经销商实地调研",       "海外组", "🔴高", 6,  4,  "C00000"),
    ("2.3 东南亚市场渠道调研",           "海外组", "🟡中", 5,  3,  "ED7D31"),
    ("2.4 日韩市场渠道调研",             "海外组", "🟡中", 7,  2,  "ED7D31"),
    ("2.5 亚马逊/eBay卖家数据分析",      "数据组", "🔴高", 5,  4,  "C00000"),
    ("2.6 竞品定价策略分析",             "数据组", "🟡中", 6,  3,  "ED7D31"),

    # 阶段三：数据整合（6月）
    ("【阶段三】数据整合与分析",         "",     "",     9,  0,  "1F3864"),
    ("3.1 全球经销商数据库建立",         "数据组", "🔴高", 9,  3,  "C00000"),
    ("3.2 竞品SWOT分析报告",             "策略组", "🔴高", 10, 3,  "C00000"),
    ("3.3 市场空白点识别",               "策略组", "🟡中", 11, 2,  "ED7D31"),
    ("3.4 潜在合作伙伴名单整理",         "商务组", "🟡中", 10, 4,  "ED7D31"),

    # 阶段四：策略制定（7月）
    ("【阶段四】渠道策略制定",           "",     "",     13, 0,  "1F3864"),
    ("4.1 目标市场优先级排序",           "策略组", "🔴高", 13, 2,  "C00000"),
    ("4.2 经销商合作方案设计",           "商务组", "🔴高", 14, 3,  "C00000"),
    ("4.3 定价与利润模型建立",           "财务组", "🟡中", 13, 3,  "ED7D31"),
    ("4.4 品牌推广策略规划",             "市场组", "🟡中", 15, 2,  "ED7D31"),

    # 阶段五：落地执行（8月）
    ("【阶段五】渠道落地执行",           "",     "",     17, 0,  "1F3864"),
    ("5.1 重点经销商接触与洽谈",         "商务组", "🔴高", 17, 4,  "C00000"),
    ("5.2 样品寄送与测试",               "产品组", "🟡中", 17, 3,  "ED7D31"),
    ("5.3 合同模板准备",                 "法务组", "🟡中", 18, 2,  "ED7D31"),
    ("5.4 市场调研总结报告",             "策略组", "🔴高", 19, 2,  "C00000"),
    ("5.5 Q4渠道拓展计划制定",           "策略组", "🔴高", 19, 2,  "C00000"),
]

# 甘特图颜色
GANTT_COLORS = {
    "C00000": "FF9999",  # 高优先级 - 红
    "ED7D31": "FFD966",  # 中优先级 - 黄
    "375623": "A9D18E",  # 低优先级 - 绿
    "1F3864": "BDD7EE",  # 分组标题 - 蓝
}

for row_idx, (task, owner, priority, start_w, duration, color) in enumerate(tasks, 4):
    ws_gantt.row_dimensions[row_idx].height = 24

    is_group = (color == "1F3864")

    # 任务名
    c1 = ws_gantt.cell(row=row_idx, column=1, value=task)
    if is_group:
        c1.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        c1.fill = PatternFill("solid", fgColor="1F3864")
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws_gantt.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=23)
    else:
        c1.font = Font(name="微软雅黑", bold=False, size=9)
        c1.fill = PatternFill("solid", fgColor="F2F2F2")
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border = tborder()

        c2 = ws_gantt.cell(row=row_idx, column=2, value=owner)
        dstyle(c2, bg="F2F2F2", sz=9, align="center"); c2.border = tborder()

        c3 = ws_gantt.cell(row=row_idx, column=3, value=priority)
        dstyle(c3, bg="F2F2F2", sz=9, align="center"); c3.border = tborder()

        # 甘特条
        bar_color = GANTT_COLORS.get(color, "BDD7EE")
        for w in range(duration):
            col_idx = 4 + (start_w - 1) + w
            if col_idx <= 23:
                gc = ws_gantt.cell(row=row_idx, column=col_idx, value="")
                gc.fill = PatternFill("solid", fgColor=bar_color)
                gc.border = Side(style="thin", color="FFFFFF")

        # 空白格
        for col_idx in range(4, 24):
            cell = ws_gantt.cell(row=row_idx, column=col_idx)
            if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                cell.border = tborder()

# 图例
ws_gantt.row_dimensions[35].height = 30
banner(ws_gantt, 35, "图例说明", "595959", sz=10, end_col="T", height=28)
legend = [
    ("🔴 高优先级（红色）", "FF9999"),
    ("🟡 中优先级（黄色）", "FFD966"),
    ("🟢 低优先级（绿色）", "A9D18E"),
]
for i, (text, bg) in enumerate(legend, 36):
    ws_gantt.row_dimensions[i].height = 22
    c = ws_gantt.cell(row=i, column=1, value=text)
    c.font = Font(name="微软雅黑", size=9)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_gantt.merge_cells(f"A{i}:D{i}")

ws_gantt.freeze_panes = "D4"

# ══════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════
out = r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v3_20260326.xlsx"
wb.save(out)
print(f"OK: {out}")
