import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_20260325.xlsx")

DARK_BLUE  = "1F3864"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_GRN = "375623"
LIGHT_GRN  = "E2EFDA"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"

def hstyle(cell, bg=DARK_BLUE, fg=WHITE, sz=11):
    cell.font = Font(name="微软雅黑", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dstyle(cell, bg=WHITE, bold=False, align="left", sz=10, color="000000"):
    cell.font = Font(name="微软雅黑", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, text, bg, fg=WHITE, sz=13, end_col="G", height=40):
    r = ws.row_dimensions[row]
    r.height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=sz, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")

# ══════════════════════════════════════════════
# Sheet 9 — Go Outdoors 详细信息
# ══════════════════════════════════════════════
ws9 = wb.create_sheet("Go Outdoors英国")
ws9.sheet_view.showGridLines = False

# Banner
banner(ws9, 1, "Go Outdoors UK — 英国户外零售巨头公司详解 & 竞品分析", "1F497D", sz=14, end_col="H")

# 公司基本信息
binfo = [
    ("品牌名称", "Go Outdoors"),
    ("公司类型", "户外用品专业零售商（Retailer）"),
    ("成立时间", "1998年"),
    ("创始人", "Paul Caplan 和 John Graham"),
    ("总部地址", "英国谢菲尔德（Sheffield, South Yorkshire, England）"),
    ("官网", "https://www.gooutdoors.co.uk"),
    ("业务定位", "英国最大户外用品专业零售商，一站式户外装备与服饰解决方案"),
    ("主要产品", "帐篷、露营装备、户外服饰、登山鞋类、户外配件"),
    ("销售模式", "线下实体店（58家超级门店）+ 线上电商 + 全渠道（Omnichannel）"),
    ("目标客群", "户外爱好者、专业露营玩家、登山徒步人群、日常休闲消费者"),
    ("合作品牌", "The North Face、Black Diamond、Petzl、Mountain Hardwear、ARC'TERYX、Salomon等近250个知名品牌"),
]

for i, (k, v) in enumerate(binfo, 3):
    bg = LIGHT_BLUE if i % 2 == 1 else WHITE
    kc = ws9.cell(row=i, column=1, value=k)
    vc = ws9.cell(row=i, column=2, value=v)
    hstyle(kc, bg="2E75B6", sz=10)
    dstyle(kc, bg="2E75B6", bold=True, color=WHITE, align="center")
    dstyle(vc, bg=bg)
    ws9.merge_cells(f"B{i}:H{i}")
    ws9.row_dimensions[i].height = 26

# 财务数据
banner(ws9, 15, "财务数据（2015/16财年，被JD Sports收购前）", "1F497D", sz=12, end_col="H", height=36)
h_fin = ["指标", "数据", "备注"]
w_fin = [22, 20, 50]
for c, (h, w) in enumerate(zip(h_fin, w_fin), 1):
    cell = ws9.cell(row=16, column=c, value=h)
    hstyle(cell, bg="2E75B6")
    ws9.column_dimensions[get_column_letter(c)].width = w
ws9.row_dimensions[16].height = 36

fin_data = [
    ["店铺数量", "58家超级门店", "主要位于英国城郊购物中心（out-of-town retail parks）"],
    ["年销售额", "2.02亿英镑 (£202.2M)", "截止2016年1月30日财年"],
    ["营业利润", "610万英镑 (£6.1M)", "截止2016年1月30日财年"],
    ["税前利润", "490万英镑 (£4.9M)", "截止2016年1月30日财年"],
    ["资产总额", "7640万英镑 (£76.4M)", "截止2016年1月30日财年"],
]
for i, row in enumerate(fin_data, 17):
    bg = WHITE if i % 2 == 0 else LIGHT_BLUE
    for c, val in enumerate(row, 1):
        cell = ws9.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=10)
        cell.border = tborder()
    ws9.row_dimensions[i].height = 28

# 股权变更
banner(ws9, 23, "股权变更历史 & 母公司情况", "1F497D", sz=12, end_col="H", height=36)
h_history = ["时间", "事件", "金额/说明", "相关方"]
w_history = [16, 30, 30, 30]
for c, (h, w) in enumerate(zip(h_history, w_history), 1):
    cell = ws9.cell(row=24, column=c, value=h)
    hstyle(cell, bg="2E75B6")
    ws9.column_dimensions[get_column_letter(c)].width = w
ws9.row_dimensions[24].height = 36

history_data = [
    ["1998年", "公司成立", "由Paul Caplan和John Graham在谢菲尔德创立", "Paul Caplan / John Graham"],
    ["2016年前", "私募投资期", "获得YFM Equity Partners和3i Group投资支持", "YFM Equity Partners / 3i Group"],
    ["2016年11月", "JD Sports收购", "JD Sports Fashion以1.123亿英镑收购Go Outdoors", "JD Sports Fashion（英国上市公司）"],
    ["收购同时", "承担债务", "JD Sports额外承担Go Outdoors约1600万英镑净负债", "JD Sports Finance"],
    ["2025年11月", "进入破产管理", "JD Sports向法院申请Go Outdoors进入破产管理程序", "JD Sports Fashion"],
    ["2025年11月", "德勤担任管理人", "Deloitte（德勤）被任命为破产管理人处理债务重组", "Deloitte（德勤）"],
]
for i, row in enumerate(history_data, 25):
    bg = WHITE if i % 2 == 0 else LIGHT_BLUE
    alert = row[0] in ["2025年11月"]
    for c, val in enumerate(row, 1):
        cell = ws9.cell(row=i, column=c, value=val)
        color = "C00000" if alert else "000000"
        dstyle(cell, bg=bg, bold=(c==1), sz=9, color=color)
        cell.border = tborder()
    ws9.row_dimensions[i].height = 30

# 当前状态
banner(ws9, 32, "⚠️ 当前状态 & 重要提示", "C00000", sz=12, end_col="H", height=44)
ws9.row_dimensions[33].height = 70
c33 = ws9.cell(row=33, column=1,
    value="【破产管理状态】\nGo Outdoors 目前正处于破产管理程序（Administration）中，由德勤（Deloitte）管理。"
          " 这意味着：\n"
          "① 原公司债务问题已无法正常运营，债权人利益由破产管理人保护\n"
          "② 业务可能仍在持续（管理人可能寻找收购方或进行债务重组）\n"
          "③ 原有供应商/合作伙伴关系可能发生变化\n"
          "④ 如需商务联系，建议通过德勤破产管理人渠道\n\n"
          "【投资/并购机会】\n由于进入破产管理，可能是收购该品牌的窗口机会，可联系德勤洽谈收购事宜。")
c33.font = Font(name="微软雅黑", size=10, color="C00000")
c33.fill = PatternFill("solid", fgColor="FFF2CC")
c33.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws9.merge_cells("A33:H33")

# 竞争分析
banner(ws9, 35, "竞争分析（与Flextail等品牌的竞争关系）", "1F497D", sz=12, end_col="H", height=36)
h_comp = ["对比维度", "Go Outdoors", "Flextail 鱼尾", "电小二/Jackery", "正浩 EcoFlow"]
w_comp = [20, 26, 26, 26, 26]
for c, (h, w) in enumerate(zip(h_comp, w_comp), 1):
    cell = ws9.cell(row=36, column=c, value=h)
    hstyle(cell, bg="2E75B6")
    ws9.column_dimensions[get_column_letter(c)].width = w
ws9.row_dimensions[36].height = 36

comp_data = [
    ["公司类型", "零售商（Retailer）", "品牌商+制造商", "品牌商+制造商", "品牌商+制造商"],
    ["核心业务", "户外装备零售", "轻量化户外电器", "户外储能电源", "移动储能系统"],
    ["成立时间", "1998年", "2015年", "2012年", "2017年"],
    ["全球覆盖", "英国为主（58家店）", "160+国家", "全球", "全球"],
    ["年销售规模", "2.02亿英镑（2016年）", "5亿+人民币", "400万台+累计", "行业头部"],
    ["主要渠道", "线下门店+线上电商", "亚马逊/独立站/1688", "京东/天猫/亚马逊", "全球全渠道"],
    ["合作可能", "★★（破产管理中）", "★★★★★（直接竞品）", "★★★★★（参考标杆）", "★★★★★（参考标杆）"],
    [" Flextail关系", "非直接竞争（零售vs品牌）", "—", "—", "—"],
]
for i, row in enumerate(comp_data, 37):
    bg = WHITE if i % 2 == 0 else LIGHT_BLUE
    for c, val in enumerate(row, 1):
        cell = ws9.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=9)
        cell.border = tborder()
    ws9.row_dimensions[i].height = 30

# 联系信息
banner(ws9, 46, "联系信息 & 参考链接", "1F497D", sz=12, end_col="H", height=36)
h_con = ["渠道", "内容", "备注"]
w_con = [20, 40, 36]
for c, (h, w) in enumerate(zip(h_con, w_con), 1):
    cell = ws9.cell(row=47, column=c, value=h)
    hstyle(cell, bg="2E75B6")
    ws9.column_dimensions[get_column_letter(c)].width = w
ws9.row_dimensions[47].height = 36

con_data = [
    ["官方网站", "https://www.gooutdoors.co.uk", "英国最大户外零售官网，可查看在售品牌和产品"],
    ["Ultimate Outdoors（旗下）", "https://www.ultimateoutdoors.co.uk", "Go Outdoors集团旗下户外零售网站"],
    ["JD Sports（母公司）", "https://www.jdsports.co.uk", "英国运动时尚零售巨头，JD Sports为上市公司"],
    ["破产管理人", "Deloitte 德勤", "可通过德勤官网联系破产管理部门洽谈收购或债权事宜"],
    ["华丽志报道", "https://luxe.co/post/50483", "2016年JD Sports收购Go Outdoors的详细报道"],
    ["FashionNetwork", "https://us.fashionnetwork.com", "JD Sports收购案时尚行业报道来源"],
    ["Retail News UK", "https://apparelresources.com", "JD-owned Go Outdoors进入破产管理新闻"],
]

# 设置A列宽度
ws9.column_dimensions["A"].width = 20
ws9.column_dimensions["B"].width = 42
ws9.column_dimensions["C"].width = 38
ws9.column_dimensions["D"].width = 30
ws9.column_dimensions["E"].width = 26
ws9.column_dimensions["F"].width = 26
ws9.column_dimensions["G"].width = 26
ws9.column_dimensions["H"].width = 26

for i, row in enumerate(con_data, 48):
    bg = WHITE if i % 2 == 0 else LIGHT_BLUE
    for c, val in enumerate(row, 1):
        cell = ws9.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=9)
        cell.border = tborder()
    ws9.row_dimensions[i].height = 28

ws9.freeze_panes = "A2"

# ══════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════
out = r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_20260325.xlsx"
wb.save(out)
print(f"OK: {out}")
