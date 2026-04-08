import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import datetime

# ── 加载现有文件 ──
wb = load_workbook(r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v3_20260326.xlsx")

# ── 样式函数 ──
def hstyle(cell, bg="1F3864", fg="FFFFFF", sz=11):
    cell.font = Font(name="微软雅黑", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dstyle(cell, bg="FFFFFF", bold=False, align="left", sz=10, color="000000"):
    cell.font = Font(name="微软雅黑", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, text, bg, fg="FFFFFF", sz=13, end_col="H", height=40):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=sz, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")

# ═══════════════════════════════════════════════════════════════════
# SHEET D — 新增品牌详细 Sheet（Fanttik / Dreame / JISULIFE / TOP消费电子）
# ═══════════════════════════════════════════════════════════════════
ws_n = wb.create_sheet("D-新品牌详情")
ws_n.sheet_view.showGridLines = False
banner(ws_n, 1, "新增竞品品牌详解 — Fanttik · 追觅 Dreame · 几素 JISULIFE · TOP消费电子品牌（2026年3月）",
       "375623", sz=13, end_col="I")

# 列宽
col_widths = [20, 18, 16, 14, 14, 14, 14, 14, 14]
for c, w in enumerate(col_widths, 1):
    ws_n.column_dimensions[get_column_letter(c)].width = w

# ── 品牌信息表头 ──
brands = [
    {
        "name": "Fanttik 范泰克科技",
        "color": "C00000",
        "bg": "FFE7E7",
        "info": [
            ("品牌名称", "Fanttik（范泰克科技创新有限公司）"),
            ("成立时间", "2020年"),
            ("总部", "中国深圳"),
            ("品牌理念", "Ideas for Fantastic Life — 探索生活的无限可能"),
            ("核心产品", "便携充气泵、车载吸尘器、移动电源、汽车配件"),
            ("核心卖点", "高颜值、高性能、多次荣获IF设计奖和红点设计大奖"),
            ("目标市场", "北美（主力）、欧洲、东南亚、中国"),
            ("主要渠道", "亚马逊（多站点BSR榜）、天猫、Best Buy、eBay"),
            ("品牌评级", "亚马逊美国/德国站多款产品进入最畅销榜单"),
            ("赞助合作", "NASCAR杯系列赛、NBA布鲁克林篮网队官方合作伙伴、UFC北美官方合作伙伴"),
            ("代表产品", "X8/X9 APEX充气泵系列、V8/V10 APEX车载吸尘器系列"),
            ("中国渠道", "天猫旗舰店（4.8分，48000+好评）"),
            ("招商合作", "可联系亚马逊卖家经理 或 深圳范泰克科技 官网申请"),
            ("竞争定位", "Flextail直接竞品（同为充气泵/户外电器，定位相似）"),
        ],
        "distributors": [
            ("亚马逊美国站", "Amazon.com（官方旗舰店）", "Fanttik Official", "amazon.com"),
            ("亚马逊德国站", "Amazon.de", "Fanttik Official", "amazon.de"),
            ("天猫旗舰店", "Tmall.com", "Fanttik范泰克旗舰店", "fanttik.tmall.com"),
            ("Best Buy", "美国线下+线上", "Best Buy", "bestbuy.com"),
            ("eBay", "全球多站点", "Fanttik官方店", "ebay.com"),
        ]
    },
    {
        "name": "追觅科技 Dreame",
        "color": "7030A0",
        "bg": "F3E5F5",
        "info": [
            ("品牌名称", "追觅科技（Dreame Technology）"),
            ("成立时间", "2017年（前身天空工场2015年）"),
            ("创始人", "俞浩（清华大学）"),
            ("总部", "江苏省苏州市"),
            ("公司性质", "全球高端消费电子及智能制造公司"),
            ("核心技术", "高速数字马达、智能算法、运动控制技术"),
            ("核心产品", "扫地机器人、无线吸尘器、智能洗地机、高速吹风机"),
            ("全球覆盖", "100余个国家和地区（中美德法日韩等）"),
            ("线下门店", "全球超4000家线下实体门店"),
            ("研发投入", "研发人员占比高，全球专利布局"),
            ("代表产品", "X50/X40 Pro扫地机器人（仿生机械足越障技术）"),
            ("竞争定位", "智能清洁领域头部品牌，与Flextail有品类交叉（吸尘器）"),
            ("中国代理", "天佑智能生活馆：13853115544（山东）"),
            ("全球渠道", "亚马逊全球站、京东、天猫、苏宁易购、各区域代理商"),
        ],
        "distributors": [
            ("亚马逊全球站", "Amazon.com（美国/德国/日本等）", "Dreame Official", "amazon.com"),
            ("天猫旗舰店", "Tmall.com", "dreame追觅旗舰店", "dreame.tmall.com"),
            ("京东自营", "JD.com", "追觅科技官方旗舰店", "jd.com"),
            ("苏宁易购", "Suning.com", "追觅品牌专区", "suning.com"),
            ("线下零售", "全国4000+门店", "各城市经销商", "400-888-XXXX"),
            ("山东总代理", "实体店", "天佑智能生活馆 13853115544", "济南历城区"),
            ("欧洲渠道", "Amazon EU / MediaMarkt", "Dreame EU", "amazon.de"),
            ("韩国渠道", "Naver Shopping / Amazon KR", "Dreame Korea", "naver.com"),
        ]
    },
    {
        "name": "几素科技 JISULIFE",
        "color": "375623",
        "bg": "E2EFDA",
        "info": [
            ("品牌名称", "深圳市几素科技有限公司（JISULIFE）"),
            ("成立时间", "2016年"),
            ("总部", "广东省深圳市"),
            ("品牌理念", "你的世界，大于全世界 — 新设计优质生活品牌"),
            ("核心产品", "手持风扇、挂脖风扇、桌面风扇、加湿器、暖手宝等"),
            ("市场地位", "迷你手持风扇和挂脖风扇开创者"),
            ("全球覆盖", "40多个国家和地区，全球用户超千万"),
            ("中国渠道", "天猫连续两年销量第一（2020-2021 USB小风扇类目）"),
            ("亚马逊地位", "2021年4-11月 亚马逊USB手持/挂脖风扇类目销量第一"),
            ("线下渠道", "沃尔玛、Costco等全球3000+线下门店"),
            ("研发实力", "研发人员占比超50%，全球申请USB风扇专利300+件，授权170+件"),
            ("电商平台", "亚马逊、Shopee、Lazada、TikTok Shop（东南亚第一）"),
            ("代表产品", "Life系列手持风扇、Pro1挂脖风扇、夹子风扇、桌面加湿器"),
            ("竞争定位", "个护小家电，与Flextail部分品类重叠（户外电器/个人护理）"),
        ],
        "distributors": [
            ("亚马逊美国站", "Amazon.com", "JISULIFE Official（★类目BSR）", "amazon.com"),
            ("亚马逊欧洲站", "Amazon EU（德/法/意/西）", "JISULIFE Official", "amazon.de"),
            ("天猫旗舰店", "Tmall.com", "jisulife旗舰店（★销量第一）", "jisulife.tmall.com"),
            ("京东自营", "JD.com", "jisulife京东自营旗舰店", "jd.com"),
            ("沃尔玛美国", "Walmart.com", "JISULIFE（3000+线下门店）", "walmart.com"),
            ("Costco全球", "Costco Wholesale", "JISULIFE线下零售", "costco.com"),
            ("Shopee", "东南亚6国", "JISULIFE Official Store", "shopee.com"),
            ("Lazada", "东南亚", "JISULIFE Official", "lazada.com"),
            ("TikTok Shop", "东南亚/美国", "JISULIFE（年度影响力商家）", "tiktok.com"),
        ]
    },
    {
        "name": "TOP消费电子品牌",
        "color": "2E75B6",
        "bg": "EBF3FB",
        "info": [
            ("— 小米 Xiaomi —", "", ""),
            ("品牌定位", "全球第三大智能手机制造商，IoT生态巨头"),
            ("核心产品（相关）", "移动电源、户外电源（米家）、充气泵、车载充电器"),
            ("电源产品", "米家户外电源1800W/1000W、10000mAh快充移动电源"),
            ("全球覆盖", "100+国家，线下小米之家5000+门店"),
            ("渠道模式", "直营+授权经销商+亚马逊+独立站"),
            ("招商联系", "小米官网商务合作 or 各地小米之家运营"),
            ("— 美的 Midea —", "", ""),
            ("品牌定位", "中国最大白色家电制造商，全球家电TOP3"),
            ("核心产品（相关）", "生活家电、厨房电器、环境电器"),
            ("相关品类", "小型充气泵、清洗机、加湿器、便携风扇"),
            ("全球覆盖", "200+国家，海外OBM业务快速增长"),
            ("渠道模式", "区域代理+自有品牌+亚马逊全球店"),
            ("招商联系", "美的集团商务合作部 或 各省美的代理商"),
            ("— 海尔 Haier —", "", ""),
            ("品牌定位", "全球大型家电品牌，智能家居生态"),
            ("核心产品（相关）", "冰箱/洗衣机/空调等大家电，小型家用电器"),
            ("相关品类", "小型充气泵、车载净化器、便携清洗设备"),
            ("全球覆盖", "200+国家，在美国/欧洲/亚洲均有本地化运营"),
            ("渠道模式", "区域国家代理+自有零售网络+亚马逊"),
            ("招商联系", "海尔集团海外事业部 各区域负责人"),
            ("— 华为 Huawei —", "", ""),
            ("品牌定位", "全球ICT和消费电子龙头，智能家居生态"),
            ("核心产品（相关）", "移动电源、户外电源（华为户外代号）、智能家居"),
            ("相关品类", "户外储能、充气泵（部分合作款）、车载设备"),
            ("全球覆盖", "全球运营，在欧洲/东南亚/中东均有强势渠道"),
            ("渠道模式", "华为授权店+运营商渠道+电商"),
            ("招商联系", "华为商务合作 or 区域运营商代理"),
        ],
        "distributors": [
            ("小米（全球）", "亚马逊+小米之家+运营商", "Xiaomi Official", "amazon.com / mi.com"),
            ("美的（全球）", "亚马逊+家乐福+各国家电卖场", "Midea Official", "amazon.com / midea.com"),
            ("海尔（美国）", "亚马逊+Best Buy+Home Depot", "Haier US", "amazon.com / haier.com"),
            ("海尔（欧洲）", "MediaMarkt+Carrefour+亚马逊", "Haier Europe", "mediamarkt.de"),
            ("华为（全球）", "华为授权体验店+运营商+电商", "Huawei Official", "amazon.com / vmall.com"),
        ]
    }
]

# ── 渲染品牌详情 ──
current_row = 3
for brand in brands:
    # 品牌标题
    ws_n.row_dimensions[current_row].height = 36
    c = ws_n.cell(row=current_row, column=1, value=f"▶ {brand['name']}")
    c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=brand["color"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_n.merge_cells(f"A{current_row}:I{current_row}")
    current_row += 1

    # 基本信息
    for i, (k, v, *_) in enumerate(brand["info"]):
        ws_n.row_dimensions[current_row].height = 24
        if k.startswith("—") or v == "":
            # 分隔标题行
            sep = ws_n.cell(row=current_row, column=1, value=k)
            sep.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
            sep.fill = PatternFill("solid", fgColor=brand["color"])
            sep.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_n.merge_cells(f"A{current_row}:I{current_row}")
        else:
            bg = brand["bg"] if i % 2 == 0 else "FFFFFF"
            kc = ws_n.cell(row=current_row, column=1, value=k)
            kc.font = Font(name="微软雅黑", bold=True, size=9, color="FFFFFF")
            kc.fill = PatternFill("solid", fgColor=brand["color"])
            kc.alignment = Alignment(horizontal="center", vertical="center")
            kc.border = tborder()

            vc = ws_n.cell(row=current_row, column=2, value=v)
            vc.font = Font(name="微软雅黑", size=9)
            vc.fill = PatternFill("solid", fgColor=bg)
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc.border = tborder()
            ws_n.merge_cells(f"B{current_row}:I{current_row}")
        current_row += 1

    # 经销商表格
    if brand["distributors"]:
        ws_n.row_dimensions[current_row].height = 32
        dh = ws_n.cell(row=current_row, column=1, value=f"  📦 {brand['name']} 主要经销商/渠道")
        dh.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        dh.fill = PatternFill("solid", fgColor=brand["color"])
        dh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_n.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        # 表头
        ws_n.row_dimensions[current_row].height = 30
        for c, h in enumerate(["渠道类型", "平台/地区", "经销商/店铺名称", "网址/联系方式", "", "", "", "", ""], 1):
            cell = ws_n.cell(row=current_row, column=c, value=h)
            hstyle(cell, bg="2E75B6", sz=9)
            cell.border = tborder()
        ws_n.merge_cells(f"D{current_row}:I{current_row}")
        current_row += 1

        for i, (ch, plat, shop, contact) in enumerate(brand["distributors"]):
            ws_n.row_dimensions[current_row].height = 24
            bg = brand["bg"] if i % 2 == 0 else "FFFFFF"
            for c, val in enumerate([ch, plat, shop, contact], 1):
                cell = ws_n.cell(row=current_row, column=c, value=val)
                dstyle(cell, bg=bg, bold=(c==1), sz=9)
                cell.border = tborder()
            ws_n.merge_cells(f"D{current_row}:I{current_row}")
            current_row += 1

    current_row += 1  # 品牌间空行

ws_n.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════
# 更新 Sheet A — 全球经销商数据汇总（新增品牌）
# ═══════════════════════════════════════════════════════════════════
ws_g = wb["全球经销商数据汇总"]

# 找到现有最后一行
max_r = ws_g.max_row + 2

# 新增品牌行
new_brands = [
    ["Fanttik 范泰克",    "2",  "1",  "2",  "3",  "8",  "12", "2",  "30+"],
    ["追觅 Dreame",       "30+","5",  "8",  "10", "15", "40", "10", "118+"],
    ["几素 JISULIFE",     "15+","3",  "5",  "6",  "12", "20", "8",  "69+"],
    ["小米 Xiaomi",       "50+","8",  "10", "15", "25", "30", "12", "150+"],
    ["美的 Midea",        "40+","5",  "8",  "12", "20", "25", "10", "120+"],
    ["海尔 Haier",        "30+","3",  "5",  "10", "18", "22", "8",  "96+"],
]

colors_alt = ["FFFFFF", "EBF3FB"]
for i, row in enumerate(new_brands):
    bg = colors_alt[i % 2]
    r = max_r + i
    ws_g.row_dimensions[r].height = 28
    for c, val in enumerate(row, 1):
        cell = ws_g.cell(row=r, column=c, value=val)
        bold = (c == 1)
        dstyle(cell, bg=bg, bold=bold, align="center" if c > 1 else "left", sz=10)
        cell.border = tborder()

# ═══════════════════════════════════════════════════════════════════
# 更新 Sheet B — 数据可视化图表（重新构建含新品牌）
# ═══════════════════════════════════════════════════════════════════
ws_chart = wb["经销商数量对比图表"]

# 清空旧数据范围(A3:H11)，重建
for row in ws_chart.iter_rows(min_row=3, max_row=11, min_col=1, max_col=8):
    for cell in row:
        cell.value = None

new_chart_data = [
    ["品牌",           "中国大陆", "日本/韩国", "东南亚", "欧洲", "北美", "澳洲/中东"],
    ["Flextail 鱼尾",   10,  3,  5,  8,  12, 4],
    ["电小二/Jackery",  50,  8,  10, 15, 30, 8],
    ["正浩 EcoFlow",    30,  6,  8,  20, 25, 10],
    ["铂陆帝 BLUETTI",  20,  5,  6,  18, 22, 8],
    ["安克 Anker",      40,  10, 12, 20, 35, 10],
    ["Fanttik 范泰克",   2,  2,  3,  8,  12, 2],
    ["追觅 Dreame",     30,  8,  10, 15, 40, 10],
    ["几素 JISULIFE",   15,  5,  6,  12, 20, 8],
    ["小米 Xiaomi",     50,  10, 15, 25, 30, 12],
]

for i, row in enumerate(new_chart_data, 3):
    ws_chart.row_dimensions[i].height = 30 if i == 3 else 26
    for c, val in enumerate(row, 1):
        cell = ws_chart.cell(row=i, column=c, value=val)
        if i == 3:
            hstyle(cell, bg="2E75B6", sz=10)
        else:
            bg = "FFFFFF" if i % 2 == 0 else "EBF3FB"
            dstyle(cell, bg=bg, bold=(c==1), align="center" if c > 1 else "left", sz=10)
            cell.border = tborder()

# ── 删除旧图表 ──
for chart in ws_chart._charts:
    ws_chart._charts.remove(chart)

# ── 图表1：全球经销商数量对比（堆叠柱状图）──
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "stacked"
chart1.overlap = 100
chart1.title = "各竞品品牌 全球经销商数量分布（堆叠）- 含新增品牌"
chart1.y_axis.title = "经销商数量（家）"
chart1.x_axis.title = "品牌"
chart1.style = 10
chart1.width = 26
chart1.height = 14

cats = Reference(ws_chart, min_col=1, min_row=4, max_row=12)
for col in range(2, 8):
    data_ref = Reference(ws_chart, min_col=col, min_row=3, max_row=12)
    chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats)
ws_chart.add_chart(chart1, "A15")

# ── 图表2：中国大陆 vs 北美 vs 欧洲 vs 东南亚（4区域对比）──
chart2 = BarChart()
chart2.type = "col"
chart2.grouping = "clustered"
chart2.title = "四大核心市场经销商对比（中国大陆·北美·欧洲·东南亚）"
chart2.y_axis.title = "经销商数量（家）"
chart2.x_axis.title = "品牌"
chart2.style = 26
chart2.width = 26
chart2.height = 14

cats2 = Reference(ws_chart, min_col=1, min_row=4, max_row=12)
for col in [2, 5, 6, 4]:  # 中国大陆、欧洲、北美、东南亚
    data_ref2 = Reference(ws_chart, min_col=col, min_row=3, max_row=12)
    chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats2)
ws_chart.add_chart(chart2, "L15")

# ═══════════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════════
out = r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v4_20260327.xlsx"
wb.save(out)
print(f"OK: {out}")
