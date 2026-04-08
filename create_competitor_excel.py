import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 颜色定义 ──
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
ACCENT_RED  = "C00000"
ACCENT_ORG  = "ED7D31"
ACCENT_GRN  = "70AD47"
LIGHT_GREY  = "F2F2F2"
WHITE       = "FFFFFF"

def header_style(cell, bg=DARK_BLUE, fg=WHITE, bold=True, size=11):
    cell.font = Font(name="微软雅黑", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def data_style(cell, bg=WHITE, bold=False, align="left"):
    cell.font = Font(name="微软雅黑", bold=bold, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)

# ══════════════════════════════════════════════
# Sheet 1 — Flextail 竞品分析
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Flextail竞品分析"
ws1.sheet_view.showGridLines = False

headers1 = ["品牌名称", "英文名", "所属公司", "成立时间", "核心产品", "目标市场", "主要渠道", "年销售额", "优势", "劣势", "官网/参考链接"]
col_widths1 = [18, 16, 22, 12, 28, 20, 20, 14, 28, 28, 36]

r = ws1.row_dimensions[1]
r.height = 36
for c, (h, w) in enumerate(zip(headers1, col_widths1), 1):
    cell = ws1.cell(row=1, column=c, value=h)
    header_style(cell)
    ws1.column_dimensions[get_column_letter(c)].width = w

data1 = [
    ["电小二", "Dxpower", "广东电小二科技有限公司", "2014年", "户外电源(80-3000W)、太阳能板", "全球户外露营/应急用电", "京东/天猫/亚马逊", "行业领先", "行业开创者，行业标准起草单位，品牌知名度高", "价格较高，产品线相对单一", "https://www.dxpower.com.cn"],
    ["正浩 EcoFlow", "EcoFlow", "深圳市正浩创新科技股份有限公司", "2017年", "移动储能、智能户外空调、车载冰箱", "全球家庭应急/户外/专业作业", "线上+线下全球渠道", "全球便携储能头部品牌", "技术领先，产品线丰富，融资能力强", "产品定价偏高", "https://www.ecoflow.com"],
    ["纽曼", "Newsmy", "纽曼数码科技有限公司", "1996年", "户外电源、汽车启动电源、GPS等", "国内3C数码/汽车电子", "线下渠道+电商", "国内老牌，品牌认知度广", "非专注户外电源领域，产品专业度有限", "http://www.newsmy.com"],
    ["公牛", "BULL", "公牛集团股份有限公司", "1995年", "户外电源、插座、开关、充电器", "家庭/办公用电", "线下渠道+电商", "渠道网络完善，品牌信任度高，安全性强", "户外产品线较新，专业性弱", "https://www.bull.cn"],
    ["安克 Anker", "Anker", "安克创新科技股份有限公司", "2011年", "移动电源、充电器、数据线、氮化镓", "全球数码配件市场", "亚马逊+独立站+线下", "上市公司，全球品牌运营能力强", "户外电源非核心业务", "https://www.anker.com"],
    ["铂陆帝 BLUETTI", "BLUETTI", "深圳市德兰明海科技有限公司", "2009年", "便携储能电源、太阳能系统", "全球户外/家庭应急", "独立站+亚马逊+线下", "产品线覆盖广，有自有工厂", "品牌知名度在国内相对较低", "https://www.bluetti-power.com"],
    ["铂陆帝(对比用)", "BLUETTI", "深圳市德兰明海科技有限公司", "2019年推出品牌", "AC30等5款户外电源", "户外野营/摄影爱好者", "京东/天猫/亚马逊", "上市公司背景", "产品型号较少", "https://baike.baidu.com/item/铂陆帝户外电源"],
    ["Jackery", "Jackery", "深圳市华宝新能源股份有限公司", "2012年", "便携太阳能发电系统、储能电源", "全球户外/家庭备电", "亚马逊+独立站", "全球便携太阳能发电系统头部品牌", "国内渠道相对薄弱", "https://www.jackery.com"],
    ["Goal Zero", "Goal Zero", "美国Goal Zero公司", "2007年", "太阳能发电机、便携电源、太阳能板", "北美户外/应急市场", "线下零售+线上", "户外电源先驱品牌", "国内渠道几乎无布局", "https://www.goalzero.com"],
    ["PowerSolar/其他品牌", "Various", "多家中小企业", "近年", "低价户外电源/贴牌产品", "价格敏感市场", "拼多多/闲鱼/1688", "价格极低", "品质无保障，无品牌力", "1688/拼多多平台"],
]

fill_colors = [WHITE, LIGHT_GREY]
for i, row in enumerate(data1, 2):
    bg = fill_colors[i % 2]
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=i, column=c, value=val)
        data_style(cell, bg=bg)
        cell.border = thin_border()

ws1.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 2 — Flextail 分销商/经销商名单
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Flextail经销商")
ws2.sheet_view.showGridLines = False

headers2 = ["公司/店铺名称", "类型", "所在地区", "主营业务/平台", "联系方式来源", "合作产品", "备注"]
col_widths2 = [30, 12, 14, 24, 20, 24, 28]

r2 = ws2.row_dimensions[1]
r2.height = 36
for c, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=c, value=h)
    header_style(cell, bg=MED_BLUE)
    ws2.column_dimensions[get_column_letter(c)].width = w

data2 = [
    ["宁波鱼尾科技有限公司", "官方/生产", "浙江宁波", "官方旗舰店/1688直营", "阿里巴巴店铺档案", "全系列 Flextail 充气泵", "flextail官方授权,综合评价高"],
    ["大连鱼尾科技有限公司(FLEXTAIL CO.)", "官方/研发", "辽宁大连", "品牌运营/产品研发", "百度百科", "户外便携式充气泵/头灯", "成立于2015年,品牌创始公司"],
    ["金华市金东区荒野户外用品商行", "分销商", "浙江金华", "1688批发/零售", "阿里巴巴实名认证", "MP2充气泵等", "10年诚信通会员"],
    ["义乌市森悦户外用品有限公司", "分销商", "浙江义乌", "1688批发/跨境", "阿里巴巴实名认证", "户外驱蚊器/充气泵", "TOP销量,跨境出口专供"],
    ["浙江森欢工贸有限公司", "分销商", "浙江", "1688/天猫", "阿里巴巴/天猫森欢户外专营店", "营地灯/驱蚊灯/充气泵", "天猫正品店,综合评分4.98"],
    ["义乌市虎盛户外用品有限公司", "分销商", "浙江义乌", "1688批发", "阿里巴巴实名认证", "真空收纳袋/适配配件", "跨境出口专供货源"],
    ["南昌市西湖区统典贸易商行", "分销商", "江西南昌", "1688批发", "阿里巴巴实名认证", "X20充气泵/MP2PLUS", "诚信通会员"],
    ["北京窝居户外", "零售商", "河北廊坊", "淘宝店", "电商平台展示", "便携驱蚊器/充气泵", "实体地址廊坊,好评率99.1%"],
    ["森欢户外专营店(天猫)", "零售商", "浙江杭州", "天猫旗舰店", "天猫平台", "充气泵/打气筒", "天猫正品店,综合4.98分"],
    ["宁波鱼尾科技1688官方店", "官方线上", "浙江宁波", "阿里巴巴", "https://detail.1688.com/offer/743620061139.html", "ZERO PUMP/全系列充气泵", "官方直营,TOP1销量"],
    ["FLEXTAIL鱼尾诚招经销商(官方)", "招商", "全国", "线上+线下+跨境", "阿里巴巴招商信息", "全系列户外产品", "禁止闲鱼/拼多多销售,有区域保护"],
]

for i, row in enumerate(data2, 2):
    bg = fill_colors[i % 2]
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=c, value=val)
        data_style(cell, bg=bg)
        cell.border = thin_border()
    ws2.row_dimensions[i].height = 30

ws2.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 3 — Vollyc 竞品分析
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("Vollyc竞品分析")
ws3.sheet_view.showGridLines = False

headers3 = ["品牌名称", "英文名", "所属公司", "成立时间", "核心产品", "目标市场", "主要渠道", "年销售额", "优势", "劣势", "官网/参考链接"]
col_widths3 = [18, 16, 22, 12, 28, 20, 20, 14, 28, 28, 36]

r3 = ws3.row_dimensions[1]
r3.height = 36
for c, (h, w) in enumerate(zip(headers3, col_widths3), 1):
    cell = ws3.cell(row=1, column=c, value=h)
    header_style(cell, bg=ACCENT_ORG)
    ws3.column_dimensions[get_column_letter(c)].width = w

data3 = [
    ["F Flextail", "Flextail/FLEXTAIL", "大连/宁波鱼尾科技", "2015年", "户外充气泵/驱蚊器/营地灯", "全球户外/跨境电商", "亚马逊/独立站/1688/线下", "5亿元+人民币", "全球160+国家,轻量化技术领先,年销5亿+", "产品线较单一(充气泵为核心)", "https://www.flextail.com"],
    ["电小二", "Dxpower", "广东电小二科技有限公司", "2014年", "户外电源(80-3000W)", "全球户外/应急", "京东/天猫/亚马逊", "行业龙头", "行业标准制定者,品牌知名度高", "无明显价格优势", "https://www.dxpower.com.cn"],
    ["正浩 EcoFlow", "EcoFlow", "深圳市正浩创新科技", "2017年", "移动储能/车载冰箱/太阳能", "全球户外/家庭", "全球全渠道", "行业头部", "技术迭代快,融资强", "定价偏高", "https://www.ecoflow.com"],
    ["铂陆帝 BLUETTI", "BLUETTI", "深圳市德兰明海科技", "2009年(品牌2019)", "便携储能/太阳能系统", "全球户外/家庭", "独立站/亚马逊", "自有工厂,产品线全", "国内品牌认知度待提升", "https://www.bluetti-power.com"],
    ["AutoCare/米家", "Various", "各大型/小型企业", "近年", "充气泵/车载配件", "国内电商", "天猫/京东/拼多多", "价格低", "品质参差不齐", "各电商平台"],
    ["Powool", "Powool", "跨境品牌", "近年", "便携充气泵/轮胎充气机", "亚马逊北美/欧洲", "亚马逊", "价格适中", "品牌知名度低", "Amazon搜索结果"],
    ["Etenwolf", "Etenwolf", "跨境品牌", "近年", "高压充气泵(160PSI)", "亚马逊北美/欧洲", "亚马逊", "专业高压产品", "产品线单一", "Amazon搜索结果"],
    ["Pow oils/杂牌", "Various", "多家白牌工厂", "近年", "低价充气泵", "价格敏感市场", "eBay/Amazon/AliExpress", "价格极低", "无品牌/品质无保证", "eBay/Amazon平台"],
]

for i, row in enumerate(data3, 2):
    bg = fill_colors[i % 2]
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=c, value=val)
        data_style(cell, bg=bg)
        cell.border = thin_border()
ws3.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 4 — Vollyc 经销商名单 (备注说明)
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("Vollyc相关经销商")
ws4.sheet_view.showGridLines = False

headers4 = ["公司/店铺名称", "类型", "所在地区", "平台/渠道", "联系方式来源", "备注"]
col_widths4 = [32, 14, 14, 22, 24, 36]

r4 = ws4.row_dimensions[1]
r4.height = 36
for c, (h, w) in enumerate(zip(headers4, col_widths4), 1):
    cell = ws4.cell(row=1, column=c, value=h)
    header_style(cell, bg=ACCENT_ORG)
    ws4.column_dimensions[get_column_letter(c)].width = w

note_cell = ws4.cell(row=2, column=1, value="⚠️ 重要说明")
note_cell.font = Font(name="微软雅黑", bold=True, size=11, color=ACCENT_RED)
note_cell.fill = PatternFill("solid", fgColor="FFF2CC")
note_cell.alignment = Alignment(horizontal="left", vertical="center")
ws4.merge_cells("A2:F2")

note_text = ws4.cell(row=3, column=1,
    value="经多轮全网搜索(中文/英文/亚马逊/eBay)，未找到名为'Vollyc'的户外电源/充气泵品牌在国内外有明确销售渠道、官网或经销商信息。"
          " VOLLYC可能是：①新成立品牌(无网络收录) ②亚马逊/eBay平台新卖家/白牌 ③品牌名拼写变体(如VOLK/VOLY等) ④特定区域小众品牌。"
          " 建议进一步核实品牌名称是否准确，或通过亚马逊/eBay直接搜索验证。下方列出的是亚马逊充气泵类目中与VOLLYC同类的竞争对手卖家。")
note_text.font = Font(name="微软雅黑", size=10)
note_text.fill = PatternFill("solid", fgColor="FFF2CC")
note_text.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws4.merge_cells("A3:F3")
ws4.row_dimensions[3].height = 60

headers4b = ["公司/店铺名称", "类型", "所在地区", "平台/渠道", "联系方式来源", "备注"]
for c, h in enumerate(headers4b, 1):
    cell = ws4.cell(row=5, column=c, value=h)
    header_style(cell, bg=ACCENT_ORG)
ws4.row_dimensions[5].height = 36

data4 = [
    ["VOLLYC Car Tire Inflator(Amazon)", "品牌方/卖家", "未知", "Amazon.com", "Amazon商品页", "亚马逊平台上销售汽车轮胎充气泵,150PSI规格"],
    ["Powools Tire Inflator", "竞品卖家", "未知", "Amazon.com", "Amazon搜索结果", "同类便携充气泵竞争对手"],
    ["Etenwolf S1 Tire Inflator", "竞品卖家", "未知", "Amazon.com/eBay", "Amazon搜索结果", "160PSI高压充气泵,欧洲市场有布局"],
    ["AutoCare Portable Tire Inflator", "竞品卖家", "未知", "Amazon.com", "Amazon搜索结果", "低价竞争产品"],
    ["注:亚马逊/eBay上大量中国卖家销售同类充气泵产品,多为白牌/工厂直销,难以一一列举", "提示", "全国/全球", "Amazon/eBay/AliExpress", "电商平台", "建议直接在亚马逊搜索'Tire Inflator Portable'或'Vollyc'获取最新卖家列表"],
]

for i, row in enumerate(data4, 6):
    bg = fill_colors[i % 2]
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=c, value=val)
        data_style(cell, bg=bg)
        cell.border = thin_border()
    ws4.row_dimensions[i].height = 40

ws4.freeze_panes = "A6"

# ══════════════════════════════════════════════
# Sheet 5 — 竞品综合对比总结
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("竞品综合对比")
ws5.sheet_view.showGridLines = False

headers5 = ["对比维度", "Flextail(鱼尾)", "Vollyc(若为充气泵)", "电小二", "正浩EcoFlow", "铂陆帝BLUETTI"]
col_widths5 = [18, 26, 26, 24, 24, 24]

r5 = ws5.row_dimensions[1]
r5.height = 40
for c, (h, w) in enumerate(zip(headers5, col_widths5), 1):
    cell = ws5.cell(row=1, column=c, value=h)
    if c == 1:
        header_style(cell, bg=DARK_BLUE)
    else:
        header_style(cell, bg=MED_BLUE)
    ws5.column_dimensions[get_column_letter(c)].width = w

compare_data = [
    ["品牌创立时间", "2015年", "待确认", "2014年", "2017年", "2019年(母公司2009年)"],
    ["主打产品", "充气泵/驱蚊器/营地灯", "待确认", "户外电源", "移动储能/太阳能系统", "便携储能/太阳能"],
    ["核心技术", "轻量化/微型化", "待确认", "锂电池BMS系统", "快充技术/X-Stream", "MPPT太阳能技术"],
    ["全球市场覆盖", "160+国家", "待确认", "全球", "全球", "全球"],
    ["年销售额", "5亿+人民币", "待确认", "行业领先", "行业头部", "持续增长"],
    ["主要销售渠道", "亚马逊/独立站/1688", "待确认", "京东/天猫/亚马逊", "全球全渠道", "独立站/亚马逊"],
    ["国内渠道", "1688/天猫/京东", "待确认", "京东/天猫", "京东/天猫", "较弱"],
    ["跨境电商能力", "极强(5亿+)", "待确认", "强", "极强", "强"],
    ["价格区间", "中高端(充气泵)", "待确认", "中高端", "中高端", "中高端"],
    ["主要竞争定位", "轻量化户外电器专家", "待确认", "户外电源开创者", "移动储能全能型", "自有工厂/定制化强"],
    ["对标参考价值", "★★★★★(直接竞品)", "待明确", "★★★★☆(户外储能参考)", "★★★★★(储能行业标杆)", "★★★☆☆(储能参考)"],
]

for i, row in enumerate(compare_data, 2):
    for c, val in enumerate(row, 1):
        bg = WHITE if c == 1 else (LIGHT_BLUE if i % 2 == 0 else WHITE)
        bold = (c == 1)
        cell = ws5.cell(row=i, column=c, value=val)
        cell.font = Font(name="微软雅黑", bold=bold, size=10)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GREY if c == 1 else bg)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws5.row_dimensions[i].height = 32

# 添加说明行
ws5.merge_cells("A13:F13")
note = ws5.cell(row=13, column=1,
    value="注: Vollyc全网搜索未找到明确品牌信息,建议核实品牌名或直接在亚马逊/eBay搜索验证。Flextail数据来源于官方资料及电商平台公开信息。")
note.font = Font(name="微软雅黑", italic=True, size=9, color="595959")
note.fill = PatternFill("solid", fgColor="F2F2F2")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws5.freeze_panes = "A2"

# ══════════════════════════════════════════════
# 保存文件
# ══════════════════════════════════════════════
output_path = r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商分析_20260325.xlsx"
wb.save(output_path)
print(f"Excel已生成: {output_path}")
