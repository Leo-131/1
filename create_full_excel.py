
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

DARK_BLUE  = "1F3864"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_GRN = "375623"
LIGHT_GRN  = "E2EFDA"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"

def hstyle(cell, bg=DARK_BLUE, fg=WHITE, sz=11):
    cell.font = Font(name="微软雅黑", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dstyle(cell, bg=WHITE, bold=False, align="left", sz=10, color="000000"):
    cell.font = Font(name="微软雅黑", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, text, bg, fg=WHITE, sz=13, end_col="G", height=40):
    r = ws.row_dimensions[row]
    r.height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=sz, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")

# ══════════════════════════════════════════════
# Sheet 1 — 竞品分析
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "竞品分析"
ws1.sheet_view.showGridLines = False
h1 = ["品牌名称","英文名","所属公司","成立/创立","核心产品","目标市场","主要渠道","年销售规模","优势亮点","劣势","官网/参考"]
w1 = [16,16,24,12,32,20,20,14,32,28,40]
for c,(h,w) in enumerate(zip(h1,w1),1):
    cell=ws1.cell(row=1,column=c,value=h); hstyle(cell); ws1.column_dimensions[get_column_letter(c)].width=w
ws1.row_dimensions[1].height=40

data1=[
    ["电小二/Jackery","Dxpower/Jackery","广东电小二科技有限公司","2012年","户外电源(80-3000W)、太阳能板","全球户外露营/应急用电","京东/天猫/亚马逊/独立站","400万台+","行业开创者，标准起草单位，全球品牌","价格较高，产品线相对单一","https://www.jackery.com"],
    ["正浩 EcoFlow","EcoFlow","深圳市正浩创新科技股份有限公司","2017年","移动储能、智能户外空调、车载冰箱、太阳能","全球家庭应急/户外","全球全渠道","行业头部","技术领先，产品线极丰富，融资能力强","产品定价偏高","https://www.ecoflow.com"],
    ["铂陆帝 BLUETTI","BLUETTI","深圳市德兰明海新能源股份有限公司","2019年推出（母司2013年）","便携储能电源、太阳能系统、UPS应急","全球户外/家庭应急","独立站/亚马逊/天猫/京东","350万+用户，847项专利","自有工厂，全球TOP3，定制化强","品牌国内知名度相对较低","https://www.bluetti.com"],
    ["纽曼","Newsmy","纽曼数码科技有限公司","1996年","户外电源、汽车启动电源、车载MP3、GPS","国内3C数码/汽车电子","线下深度渠道+电商","行业知名","老牌，品牌认知度广，渠道深","非专注户外电源","http://www.newsmy.com"],
    ["公牛 BULL","BULL","公牛集团股份有限公司","1995年","户外电源、插座、开关、充电器","家庭/办公/户外","线下深度渠道+电商","上市公司","渠道完善，品牌信任度高，安全性强","户外产品线较新","https://www.bull.cn"],
    ["安克 Anker/Solix","Anker/Solix","安克创新科技股份有限公司","2011年","移动电源、充电器、户外电源Solix系列","全球数码配件/户外储能","亚马逊+独立站+线下","8000万+用户","上市公司，全球品牌运营能力强","户外电源非核心业务","https://www.anker.com"],
    ["Jackery（美国主体）","Jackery","华宝新能/Jackery Inc.","2012年","便携太阳能发电系统、储能电源","全球户外/家庭备电","亚马逊+独立站","全球第一","全球便携太阳能发电系统头部品牌","国内渠道相对薄弱","https://www.jackery.com"],
    ["Goal Zero","Goal Zero","Goal Zero LLC（美国）","2007年","太阳能发电机、便携电源、太阳能板","北美户外/应急市场","线下零售+线上","北美领先","户外电源先驱品牌","国内渠道几乎无布局","https://www.goalzero.com"],
    ["百克龙","BAICLON","深圳市百克龙科技有限公司","约2015年","户外移动电源、UPS电源","国内户外/工业/应急","京东/天猫/1688","持续增长","国内专业储能品牌，价格有竞争力","品牌知名度一般","https://www.baiclon.com"],
    ["普闹得/Pronat","Pronat","广州Pronat户外","近年","户外储能、应急电源","国内外户外市场","亚马逊/独立站","新兴","新兴品牌，主打性价比","品牌积累少","亚马逊搜索结果"],
]
for i,row in enumerate(data1,2):
    bg=WHITE if i%2==0 else LIGHT_GREY
    for c,val in enumerate(row,1):
        cell=ws1.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c==1),sz=9); cell.border=tborder()
    ws1.row_dimensions[i].height=32
ws1.freeze_panes="A2"

# ══════════════════════════════════════════════
# Sheet 2 — Flextail 品牌详解+经销商网络
# ══════════════════════════════════════════════
ws2=wb.create_sheet("Flextail经销商")
ws2.sheet_view.showGridLines=False
banner(ws2,1,"Flextail 鱼尾科技 — 品牌详解 & 经销商网络",DARK_BLUE,sz=14,end_col="H")

brand=[
    ("品牌名称","Flextail 鱼尾 / FLEXTAIL"),
    ("成立时间","2015年"),
    ("所属公司","大连鱼尾科技有限公司（研发）+ 宁波鱼尾科技有限公司（运营/生产）"),
    ("总部地点","辽宁大连 / 浙江宁波"),
    ("核心产品","户外便携充气泵（MAX PUMP/TINY PUMP/ZERO PUMP）、户外驱蚊器、营地灯、淋浴器"),
    ("主打技术","微型化轻量化充气泵，业界最轻最小"),
    ("全球市场","160+个国家，年销售额突破5亿元人民币"),
    ("主打渠道","亚马逊（全球）、独立站（flextail.com）、1688（分销）、天猫/京东（国内）"),
    ("招商政策","线上平台+线下实体店+跨境外贸经销商；禁止闲鱼/拼多多销售；部分区域包销保护"),
    ("竞争定位","轻量化户外电器专家，充气泵品类全球领先"),
]
for i,(k,v) in enumerate(brand,3):
    bg=LIGHT_BLUE if i%2==1 else WHITE
    kc=ws2.cell(row=i,column=1,value=k); vc=ws2.cell(row=i,column=2,value=v)
    hstyle(kc,bg=MED_BLUE,sz=10); dstyle(kc,bg=MED_BLUE,bold=True,color=WHITE,align="center")
    dstyle(vc,bg=bg); ws2.merge_cells(f"B{i}:H{i}"); ws2.row_dimensions[i].height=26

banner(ws2,14,"Flextail 经销商 / 分销商 详细名单",MED_BLUE,sz=12,end_col="H",height=40)
h2=["公司/店铺名称","类型","所在地区","平台/渠道","合作产品","联系方式","备注","参考链接"]
w2=[30,12,14,22,24,26,28,38]
for c,(h,w) in enumerate(zip(h2,w2),1):
    cell=ws2.cell(row=15,column=c,value=h); hstyle(cell,bg=MED_BLUE); ws2.column_dimensions[get_column_letter(c)].width=w
ws2.row_dimensions[15].height=38

dist=[
    ["宁波鱼尾科技有限公司","官方/生产","浙江宁波","1688官方店+天猫+京东","全系列充气泵","阿里巴巴店铺实名认证","官方授权核心公司,TOP1销量","https://detail.1688.com/offer/743620061139.html"],
    ["大连鱼尾科技有限公司","官方/研发","辽宁大连","品牌运营/产品研发","户外便携充气泵/头灯","百度百科收录","品牌创始公司，2015年","https://baike.baidu.com/item/大连鱼尾科技有限公司"],
    ["金华市金东区荒野户外用品商行","分销商","浙江金华","1688批发/零售","MP2充气泵系列","阿里巴巴10年诚信通","专注户外装备分销","https://detail.1688.com/offer/575427529058.html"],
    ["义乌市森悦户外用品有限公司","分销/跨境","浙江义乌","1688批发/跨境出口","户外驱蚊器/充气泵","阿里巴巴实名认证","TOP1销量,跨境出口专供","https://detail.1688.com/offer/706552226808.html"],
    ["浙江森欢工贸有限公司","分销商","浙江","1688/天猫","营地灯/驱蚊灯/充气泵","天猫/阿里巴巴平台","天猫正品店,综合评分4.98","森欢户外专营店（天猫）"],
    ["义乌市虎盛户外用品有限公司","分销商","浙江义乌","1688批发/跨境","真空收纳袋/适配配件","阿里巴巴实名认证","跨境出口专供货源","https://detail.1688.com/offer/911825304047.html"],
    ["南昌市西湖区统典贸易商行","分销商","江西南昌","1688批发","X20充气泵/MP2PLUS","阿里巴巴诚信通会员","诚信通会员","https://detail.1688.com/offer/855988365298.html"],
    ["北京窝居户外","零售商","河北廊坊","淘宝店","便携驱蚊器/充气泵","电商平台展示","好评率99.1%","淘宝店：北京窝居户外"],
    ["森欢户外专营店（天猫）","零售商","浙江杭州","天猫旗舰店","充气泵/打气筒/营地灯","天猫平台","天猫正品店,综合4.98分","天猫搜索：森欢户外专营店"],
    ["FLEXTAIL官方招商处","品牌方","全国","线上+线下+跨境","全系列户外产品","官网/阿里巴巴","禁止闲鱼/拼多多,区域保护","https://detail.1688.com/offer/630697891979.html"],
]
for i,row in enumerate(dist,16):
    bg=WHITE if i%2==0 else LIGHT_GREY
    for c,val in enumerate(row,1):
        cell=ws2.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c==1),sz=9); cell.border=tborder()
    ws2.row_dimensions[i].height=30
ws2.freeze_panes="A16"

# ══════════════════════════════════════════════
# Sheet 3 — 电小二 竞品经销商详情
# ══════════════════════════════════════════════
ws3=wb.create_sheet("电小二经销商")
ws3.sheet_view.showGridLines=False
banner(ws3,1,"电小二 / Jackery — 竞品经销商/代理商详细名单",ACCENT_GRN,sz=13,end_col="G")

binfo3=[
    ("品牌名称","电小二（Jackery）"),
    ("所属公司","广东电小二科技有限公司"),
    ("成立时间","2012年（2016年推出全球首款户外便携储能电源）"),
    ("总部","广东深圳龙华区华侨城北站壹号大厦38楼"),
    ("核心产品","户外电源（80-3000W）、太阳能板"),
    ("全球销量","累计400万台+"),
    ("主要渠道","京东/天猫/亚马逊/独立站/线下"),
    ("全国热线","400-668-9293（售后服务）"),
    ("代理合作","苏经理 13714300193 | 褚经理 17861515971 | 叶经理 18605613145（东南亚）"),
    ("品牌合作邮箱","zhuhl@dx2.cn"),
]
for i,(k,v) in enumerate(binfo3,3):
    bg=LIGHT_GRN if i%2==1 else WHITE
    kc=ws3.cell(row=i,column=1,value=k); vc=ws3.cell(row=i,column=2,value=v)
    hstyle(kc,bg=ACCENT_GRN,sz=10); dstyle(kc,bg=ACCENT_GRN,bold=True,color=WHITE,align="center")
    dstyle(vc,bg=bg); ws3.merge_cells(f"B{i}:G{i}"); ws3.row_dimensions[i].height=26

banner(ws3,14,"电小二 代理商 / 经销商 / 零售商 名单",ACCENT_GRN,sz=12,end_col="G",height=40)
h3=["公司/店铺名称","类型","所在地区","平台/渠道","主营产品","联系方式","备注"]
w3=[32,12,14,22,26,30,26]
for c,(h,w) in enumerate(zip(h3,w3),1):
    cell=ws3.cell(row=15,column=c,value=h); hstyle(cell,bg=ACCENT_GRN); ws3.column_dimensions[get_column_letter(c)].width=w
ws3.row_dimensions[15].height=38

d3=[
    ["广东电小二科技有限公司（总部）","官方/品牌方","广东深圳","京东/天猫/亚马逊/独立站","全系列户外电源","400-668-9293 | dx2brand@dx2.cn","品牌总部，累计销量400万台+"],
    ["苏经理（全国代理合作）","品牌代理负责人","广东深圳","品牌方对接","全国代理商招募","13714300193（微信同号）","负责全国代理合作"],
    ["褚经理（全国代理合作）","品牌代理负责人","广东深圳","品牌方对接","全国代理商招募","17861515971（微信同号）","负责全国代理合作"],
    ["叶经理（东南亚代理）","品牌代理负责人","广东深圳","品牌方对接","东南亚区域代理","18605613145（微信同号）","专注东南亚市场"],
    ["山东电小二信息科技有限公司","省级代理","山东烟台","1688/天猫/充电桩","充电桩/户外电源","18660099020 | sddxr8@126.com","山东区域授权代理，8年资质"],
    ["济南恒通通讯（零售）","零售商","山东济南","中关村在线/线下实体","户外电源1000Pro2等","13006574998 / 0531-67757190","济南华强电子世界Q1039"],
    ["家家通贸易公司","贸易商","北京昌平","机电之家/电商","电小二户外电源1800W","133-9175-3024","北京区域实体供货"],
    ["Jackery电小二全国租赁网点","租赁合作","全国多省","官网/线下网点","户外电源租赁服务","400-668-9293","全国各省均有合作网点"],
]
for i,row in enumerate(d3,16):
    bg=WHITE if i%2==0 else "F0FFF0"
    for c,val in enumerate(row,1):
        cell=ws3.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c==1),sz=9); cell.border=tborder()
    ws3.row_dimensions[i].height=30
ws3.freeze_panes="A16"

# ══════════════════════════════════════════════
# Sheet 4 — 正浩EcoFlow 竞品经销商详情
# ══════════════════════════════════════════════
ws4=wb.create_sheet("正浩EcoFlow经销商")
ws4.sheet_view.showGridLines=False
banner(ws4,1,"正浩 EcoFlow — 竞品经销商/代理商/合作商详细名单","0F4C81",sz=13,end_col="G")

binfo4=[
    ("品牌名称","正浩 EcoFlow"),
    ("所属公司","深圳市正浩创新科技股份有限公司"),
    ("成立时间","2017年"),
    ("总部","广东深圳（苏州有产业园）"),
    ("核心产品","移动储能（DELTA/River系列）、智能户外空调、车载冰箱、太阳能"),
    ("主打技术","X-Stream快充技术、MPPT太阳能技术"),
    ("主要渠道","全球全渠道（独立站+亚马逊+京东+天猫+线下）"),
    ("全球热线","400-600-5753（客户服务）"),
    ("经销商合作","18923781545"),
    ("推广合作邮箱","marketing.cn@ecoflow.com"),
    ("招商官网","https://www.ecoflow.com/cn/cooperation"),
    ("全球总部","中国深圳 | 美国西雅图 | 日本东京 | 德国杜塞尔多夫"),
]
for i,(k,v) in enumerate(binfo4,3):
    bg=LIGHT_BLUE if i%2==1 else WHITE
    kc=ws4.cell(row=i,column=1,value=k); vc=ws4.cell(row=i,column=2,value=v)
    hstyle(kc,bg="0F4C81",sz=10); dstyle(kc,bg="0F4C81",bold=True,color=WHITE,align="center")
    dstyle(vc,bg=bg); ws4.merge_cells(f"B{i}:G{i}"); ws4.row_dimensions[i].height=26

banner(ws4,16,"正浩EcoFlow 代理商 / 经销商 / 分销商 名单","0F4C81",sz=12,end_col="G",height=40)
h4=["公司/店铺名称","类型","所在地区","平台/渠道","主营产品","联系方式","备注"]
w4=[34,12,14,22,26,28,26]
for c,(h,w) in enumerate(zip(h4,w4),1):
    cell=ws4.cell(row=17,column=c,value=h); hstyle(cell,bg="0F4C81"); ws4.column_dimensions[get_column_letter(c)].width=w
ws4.row_dimensions[17].height=38

d4=[
    ["深圳市正浩创新科技（总部）","官方/品牌方","广东深圳","独立站/亚马逊/京东/天猫/线下","全系列正浩产品","400-600-5753 | 18923781545","品牌总部，行业头部"],
    ["EcoFlow授权总代理（抖音渠道）","全国总代理","全国","抖音/线下渠道","全系产品+行业解决方案","抖音平台搜索","户外露营/自驾/房车/海上作业等"],
    ["山东星祺电子科技有限公司","分销商","山东","1688/顺企网","EcoFlow 1280W储能电池","1688实名认证","专注储能分销"],
    ["山东森威尔电源科技有限公司","分销商","山东","1688/顺企网","EcoFlow Delta2/Delta2Max欧规","1688实名认证","45条评价，欧洲市场专供"],
    ["广西南宁潮玩电子产品销售商行","分销商","广西南宁","顺企网/1688","EcoFlow DELTA德2/River系列","13978158226","广西区域实体供货"],
    ["广东深圳授权经销商（牛器网）","授权经销商","广东深圳","牛器网/线下","EcoFlow全系产品","400-8628-336","授权总代理"],
    ["苏州市正浩创新科技（苏州子公司）","子公司/产区","江苏苏州","恒泰智造产业园","智慧储能研发生产配套","苏州恒泰集团","2025年3月签约入驻苏州工业园"],
    ["EcoFlow美国西雅图办公室","海外总部","美国西雅图","独立站/亚马逊北美","全系产品","+1(800)368-8604 | sales@ecoflow.com","北美市场核心"],
    ["EcoFlow日本东京办公室","海外分部","日本东京","独立站/亚马逊日本","全系产品","03 6666 8366","日本及东亚市场"],
    ["EcoFlow官方招商合作页","品牌方","全球","官网申请","全品类招募授权经销商","marketing.cn@ecoflow.com","直接合作即获技术支持"],
]
for i,row in enumerate(d4,18):
    bg=WHITE if i%2==0 else LIGHT_BLUE
    for c,val in enumerate(row,1):
        cell=ws4.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c==1),sz=9); cell.border=tborder()
    ws4.row_dimensions[i].height=30
ws4.freeze_panes="A18"

# ══════════════════════════════════════════════
# Sheet 5 — 铂陆帝 BLUETTI 竞品经销商详情
# ══════════════════════════════════════════════
ws5=wb.create_sheet("铂陆帝BLUETTI经销商")
ws5.sheet_view.showGridLines=False
banner(ws5,1,"铂陆帝 BLUETTI — 竞品经销商/代理商/合作商详细名单","4A235A",sz=13,end_col="G")

binfo5=[
    ("品牌名称","铂陆帝 BLUETTI"),
    ("所属公司","深圳市德兰明海新能源股份有限公司"),
    ("成立时间","2013年（品牌2019年）"),
    ("注册资本","37987.25万元（上市公司背景）"),
    ("总部","广东深圳南山区凯达尔集团中心大厦"),
    ("核心产品","便携储能电源、太阳能系统、UPS应急电源"),
    ("全球数据","120+销售区域，847项发明专利，70000m2生产基地，350万+全球用户"),
    ("全国热线","4001-628-066"),
    ("天猫旗舰店","BLUETTI铂陆帝天猫官方旗舰店"),
    ("京东旗舰店","BLUETTI铂陆帝京东官方旗舰店"),
    ("招商政策","BLUETTI之星招募计划，全球一件代发，六大代理优势"),
    ("招商合作","https://www.bluetti.cn/sale-partner.html"),
]
for i,(k,v) in enumerate(binfo5,3):
    bg="F3E5F5" if i%2==1 else WHITE
    kc=ws5.cell(row=i,column=1,value=k); vc=ws5.cell(row=i,column=2,value=v)
    hstyle(kc,bg="4A235A",sz=10); dstyle(kc,bg="4A235A",bold=True,color=WHITE,align="center")
    dstyle(vc,bg=bg); ws5.merge_cells(f"B{i}:G{i}"); ws5.row_dimensions[i].height=26

banner(ws5,16,"铂陆帝BLUETTI 代理商 / 经销商 / 分销商 名单","4A235A",sz=12,end_col="G",height=40)
h5=["公司/店铺名称","类型","所在地区","平台/渠道","主营产品","联系方式","备注"]
w5=[34,12,14,22,26,28,26]
for c,(h,w) in enumerate(zip(h5,w5),1):
    cell=ws5.cell(row=17,column=c,value=h); hstyle(cell,bg="4A235A"); ws5.column_dimensions[get_column_letter(c)].width=w
ws5.row_dimensions[17].height=38

d5=[
    ["深圳市德兰明海新能源股份（总部）","官方/品牌方","广东深圳","独立站/亚马逊/天猫/京东","全系列BLUETTI产品","4001-628-066","上市公司，70000m2生产基地"],
    ["BLUETTI天猫旗舰店","官方直营","全国","天猫","铂陆帝全系户外电源","天猫平台搜索","官方授权直营"],
    ["BLUETTI京东旗舰店","官方直营","全国","京东","铂陆帝全系户外电源","京东平台搜索","官方授权直营"],
    ["BLUETTI官方招商合作","品牌方","全球","官网申请","全球一件代发授权经销商","https://www.bluetti.cn/sale-partner.html","6大代理优势，招募中"],
    ["上海星亿祺电子科技有限公司","分销商","上海","1688/批发","BLUETTI太阳能板120-450W","1688实名认证","50套起批，价格1409-4699元"],
    ["深圳市德兰明海1688官方店","官方/分销","广东深圳","1688官方店","BLUETTI 3072Wh储能电源","1688官方认证","官方直营，>=2个起订"],
    ["黄山铂陆帝代理商（ZOL）","区域代理","安徽黄山","中关村在线","AC180/AC60/EB70等","中关村在线平台","黄山区域授权"],
    ["全球代理商招募计划（IEAE展会）","品牌方","全球","IEAE展会/官网","全系列产品一件代发","知乎/官网","2021年IEAE广州电子展正式启动"],
]
for i,row in enumerate(d5,18):
    bg=WHITE if i%2==0 else "F3E5F5"
    for c,val in enumerate(row,1):
        cell=ws5.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c==1),sz=9); cell.border=tborder()
    ws5.row_dimensions[i].height=30
ws5.freeze_panes="A18"

# ══════════════════════════════════════════════
# Sheet 6 — 安克/公牛/其他竞品经销商
# ══════════════════════════════════════════════
ws6=wb.create_sheet("安克等其他竞品经销商")
ws6.sheet_view.showGridLines=False
banner(ws6,1,"安克 Anker / 公牛 / 纽曼 / 其他 — 竞品经销商代理商详细名单","7B4F00",sz=13,end_col="G")

h6=["公司/店铺名称","所属品牌","类型","所在地区","平台/渠道","联系方式","备注"]
w6=[34,14,14,14,22,28,26]
for c,(h,w) in enumerate(zip(h6,w6),1):
    cell=ws6.cell(row=2,column=c,value=h); hstyle(cell,bg="7B4F00"); ws6.column_dimensions[get_column_letter(c)].width=w
ws6.row_dimensions[2].height=38

d6=[
    ["安克创新科技股份有限公司（总部）","安克 Anker","官方/品牌方","湖南长沙","亚马逊+独立站+线下","400-055-0036 | 0755-33100690","上市公司，8000万+用户，100+国家"],
    ["安克天猫旗舰店","安克 Anker","官方直营","全国","天猫","天猫平台","官方授权直营"],
    ["Anker京东自营旗舰店","安克 Anker","官方直营","全国","京东","京东平台","官方授权自营"],
    ["深圳金沐雪科技有限公司","安克 Anker","分销商","广东深圳","1688","1688平台实名认证","Anker 90000mAh户外电源分销"],
    ["谭青春（安克1688分销-业务经理）","安克 Anker","分销商","广东","1688","1688平台展示","业务部经理"],
    ["公牛集团股份有限公司（总部）","公牛 BULL","官方/品牌方","浙江宁波","线下深度渠道+电商","官网联系","上市公司，渠道网络完善"],
    ["公牛户外电源官方渠道","公牛 BULL","官方直营","全国","京东/天猫/线下","官网平台","插座/开关延伸至户外储能"],
    ["纽曼数码科技有限公司（总部）","纽曼 Newsmy","官方/品牌方","湖南长沙","线下+电商","官网联系","老牌数码，1996年成立"],
    ["纽曼户外电源各电商平台","纽曼 Newsmy","零售商","全国","京东/天猫/拼多多","各电商平台","多平台均有授权店铺"],
    ["深圳市胜江源电子有限公司","多品牌分销","分销商","广东深圳","1688","1688实名认证","正浩/铂陆帝/电小二多品牌分销"],
]
for i,row in enumerate(d6,3):
    bg=WHITE if i%2==0 else "FFF8E1"
    for c,val in enumerate(row,1):
        cell=ws6.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c in [1,2]),sz=9); cell.border=tborder()
    ws6.row_dimensions[i].height=30
ws6.freeze_panes="A3"

# ══════════════════════════════════════════════
# Sheet 7 — Vollyc 说明
# ══════════════════════════════════════════════
ws7=wb.create_sheet("Vollyc说明")
ws7.sheet_view.showGridLines=False
banner(ws7,1,"Vollyc 品牌 — 全网搜索结果说明","7030A0",sz=14,end_col="E",height=50)

ws7.row_dimensions[2].height=50
c=ws7.cell(row=2,column=1,value="⚠️ 重要说明：全网未找到 Vollyc 品牌的经销商和代理商信息")
c.font=Font(name="微软雅黑",bold=True,size=12,color="7030A0")
c.fill=PatternFill("solid",fgColor="E8D5F5")
c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
ws7.merge_cells("A2:E2")

ws7.row_dimensions[3].height=90
c=ws7.cell(row=3,column=1,
    value="经多轮全网搜索（中文/英文，含百度/谷歌/亚马逊/eBay/阿里巴巴/顺企网等平台），"
          "未找到名为'Vollyc'的户外电源/充气泵品牌在国内外有明确销售渠道、官网或经销商信息。\n\n可能原因：\n"
          "① 新成立品牌（无网络收录）② 亚马逊/eBay平台新卖家/白牌/无官网\n"
          "③ 品牌名拼写变体（如VOLK/VOLY/VOLIC等）④ 特定区域小众品牌\n"
          "⑤ 可能为用户听说的品牌但实际名称不同\n\n"
          "建议：直接在亚马逊搜索'Vollyc'或'VOLLYC'验证，或提供更多背景信息（如产品类型、购买渠道等）。")
c.font=Font(name="微软雅黑",size=10)
c.fill=PatternFill("solid",fgColor="FFF2CC")
c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
ws7.merge_cells("A3:E3")

banner(ws7,5,"亚马逊/eBay 同类充气泵竞品卖家（供参考）","595959",sz=12,end_col="E",height=40)
h7=["品牌/卖家名称","类型","平台","参考链接","备注"]
w7=[28,14,16,36,28]
for c,(h,w) in enumerate(zip(h7,w7),1):
    cell=ws7.cell(row=6,column=c,value=h); hstyle(cell,bg="595959"); ws7.column_dimensions[get_column_letter(c)].width=w
ws7.row_dimensions[6].height=38

d7=[
    ["VOLLYC Car Tire Inflator（Amazon）","品牌方/卖家","Amazon.com","Amazon搜索","150PSI规格轮胎充气泵"],
    ["Powools Tire Inflator","竞品卖家","Amazon.com","Amazon搜索结果","同类便携充气泵竞争对手"],
    ["Etenwolf S1 Tire Inflator","竞品卖家","Amazon.com/eBay","Amazon搜索结果","160PSI高压充气泵，欧洲市场有布局"],
    ["AutoCare Portable Tire Inflator","竞品卖家","Amazon.com","Amazon搜索结果","低价竞争产品"],
    ["注：亚马逊/eBay大量中国卖家销售同类充气泵，多为白牌/工厂直销","提示","Amazon/eBay/AliExpress","电商平台","建议直接搜索'Tire Inflator Portable'获取最新"],
]
for i,row in enumerate(d7,7):
    bg=WHITE if i%2==0 else LIGHT_GREY
    for c,val in enumerate(row,1):
        cell=ws7.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c==1),sz=9); cell.border=tborder()
    ws7.row_dimensions[i].height=35

# ══════════════════════════════════════════════
# Sheet 8 — 品牌招商合作联系方式汇总
# ══════════════════════════════════════════════
ws8=wb.create_sheet("品牌招商联系方式汇总")
ws8.sheet_view.showGridLines=False
banner(ws8,1,"户外电源/储能行业 品牌招商合作联系方式汇总","1F3864",sz=14,end_col="F",height=50)

h8=["品牌名称","招商负责人/部门","联系电话","招商邮箱","招商网址","备注"]
w8=[18,20,22,26,38,30]
for c,(h,w) in enumerate(zip(h8,w8),1):
    cell=ws8.cell(row=2,column=c,value=h); hstyle(cell,bg="1F3864"); ws8.column_dimensions[get_column_letter(c)].width=w
ws8.row_dimensions[2].height=40

d8=[
    ["Flextail 鱼尾","官方招商处（阿里巴巴）","1688平台联系","官网联系","https://detail.1688.com/offer/630697891979.html","禁止闲鱼/拼多多，区域保护"],
    ["电小二/Jackery","苏经理/褚经理/叶经理","13714300193 / 17861515971 / 18605613145","zhuhl@dx2.cn","http://www.dx2.cn","全国代理+东南亚代理"],
    ["正浩 EcoFlow","经销商合作部","18923781545","marketing.cn@ecoflow.com","https://www.ecoflow.com/cn/cooperation","全球授权经销商招募中"],
    ["铂陆帝 BLUETTI","招商合作部","4001-628-066","官网联系","https://www.bluetti.cn/sale-partner.html","BLUETTI之星全球招募"],
    ["安克 Anker","官方渠道","400-055-0036","官网联系","https://www.anker.com","全球一件代发"],
    ["公牛 BULL","官方渠道","官网联系","官网联系","https://www.bull.cn","线下渠道完善"],
    ["Jackery","代理合作（同电小二）","400-668-9293","zhuhl@dx2.cn","https://www.jackery.com","全球户外储能领先"],
    ["Goal Zero","美国总部","官网联系","support@goalzero.com","https://www.goalzero.com","北美户外先驱"],
]
for i,row in enumerate(d8,3):
    bg=WHITE if i%2==0 else LIGHT_GREY
    for c,val in enumerate(row,1):
        cell=ws8.cell(row=i,column=c,value=val)
        dstyle(cell,bg=bg,bold=(c==1),sz=9); cell.border=tborder()
    ws8.row_dimensions[i].height=32
ws8.freeze_panes="A3"

out=r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_20260325.xlsx"
wb.save(out)
print(f"OK: {out}")
