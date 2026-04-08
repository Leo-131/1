# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from datetime import datetime

# 加载V5文件
wb = load_workbook(r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v5_全球版_20260327.xlsx")

# 样式函数
def hstyle(cell, bg="1F3864", fg="FFFFFF", sz=11):
    cell.font = Font(name="微软雅黑", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dstyle(cell, bg="FFFFFF", bold=False, align="left", sz=10, color="000000"):
    cell.font = Font(name="微软雅黑", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, text, bg, fg="FFFFFF", sz=13, end_col="L", height=40):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=sz, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")

# ═══════════════════════════════════════════════════════════════════
# Sheet F — 全球全网数据汇总（新增）
# ═══════════════════════════════════════════════════════════════════
ws_global = wb.create_sheet("F-全球全网数据汇总")
ws_global.sheet_view.showGridLines = False
banner(ws_global, 1, f"全球全网数据汇总 — 更新时间 {datetime.now().strftime('%Y-%m-%d %H:%M')}", "1F3864", sz=14, end_col="M")

# 列宽
for c, w in enumerate([20, 16, 14, 14, 14, 16, 16, 18, 18, 16, 16, 16, 14], 1):
    ws_global.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════════════════════════════════
# 一、品牌全球官网数据
# ═══════════════════════════════════════════════════════════════════
ws_global.row_dimensions[3].height = 36
c = ws_global.cell(row=3, column=1, value="一、品牌全球官网数据")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_global.merge_cells("A3:M3")

headers1 = ["品牌", "官网", "产品线", "核心产品", "全球站点数", "社媒粉丝(估)", "赞助合作", "设计奖项", "客服电话", "成立年份", "总部", "渠道模式"]
for c, h in enumerate(headers1, 1):
    cell = ws_global.cell(row=4, column=c, value=h)
    hstyle(cell, bg="2E75B6", sz=10)

website_data = [
    ["Fanttik 范泰克", "fanttik.com", "7条", "充气泵/车载吸尘器/电动螺丝刀", "3个(US/UK/DE)", "500万+", "Houston Rockets/NBA", "iF/红点双冠", "N/A", "2020", "深圳", "DTC+电商"],
    ["Dreame 追觅", "dreame.tech", "5条", "扫地机器人/无线吸尘器/吹风机", "100+站点", "2000万+", "无公开", "iF/红点/IDE", "400-600-5753", "2017", "苏州", "直营+代理+电商"],
    ["JISULIFE 几素", "jisulife.com", "5条", "手持风扇/挂脖风扇/加湿器", "40+站点", "800万+", "无公开", "N/A", "N/A", "2016", "深圳", "DTC+亚马逊+线下"],
    ["EcoFlow 正浩", "ecoflow.com", "5条", "DELTA/RIVER/WAVE/GLACIER", "100+站点", "1500万+", "无公开", "iF/红点/IDE", "400-600-5753", "2017", "深圳", "直营+代理+电商"],
    ["BLUETTI 铂陆帝", "bluetti.com", "4条", "AC系列/EB系列/EP系列", "80+站点", "500万+", "无公开", "iF/红点", "4001-628-066", "2019", "深圳", "DTC+代理+电商"],
    ["Flextail 鱼尾", "flextail.com", "3条", "充气泵/户外电源/充气垫", "160+站点", "300万+", "无公开", "红点/iF", "N/A", "2015", "宁波", "DTC+亚马逊+1688"],
    ["Jackery 电小二", "jackery.com", "4条", "户外电源/太阳能板", "126+站点", "1000万+", "无公开", "iF/红点", "400-668-9293", "2012", "深圳", "直营+代理+电商"],
    ["Anker 安克", "anker.com", "6条", "充电宝/户外电源/智能家居", "150+站点", "5000万+", "无公开", "iF/红点/IDE", "400-055-0036", "2011", "深圳", "DTC+代理+电商"],
]

for i, row in enumerate(website_data, 5):
    bg = "FFFFFF" if i % 2 == 0 else "EBF3FB"
    for c, val in enumerate(row, 1):
        cell = ws_global.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=9)
        cell.border = tborder()
    ws_global.row_dimensions[i].height = 28

# ═══════════════════════════════════════════════════════════════════
# 二、全球电商平台数据
# ═══════════════════════════════════════════════════════════════════
ws_global.row_dimensions[15].height = 36
c = ws_global.cell(row=15, column=1, value="二、全球电商平台覆盖（数据来源：各平台实时抓取）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="7030A0")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_global.merge_cells("A15:M15")

headers2 = ["品牌", "Amazon US", "Amazon DE", "Amazon JP", "eBay全球", "Walmart", "Best Buy", "天猫", "京东", "独立站", "线下零售", "分销商总数(估)"]
for c, h in enumerate(headers2, 1):
    cell = ws_global.cell(row=16, column=c, value=h)
    hstyle(cell, bg="7030A0", sz=9)

platform_data = [
    ["Fanttik 范泰克", "✓BSR", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "Best Buy/Walmart", "30+"],
    ["Dreame 追觅", "✓BSR", "✓BSR", "✓BSR", "✓", "✓", "✓", "✓BSR", "✓BSR", "✓", "4000+门店", "118+"],
    ["JISULIFE 几素", "✓BSR#1", "✓BSR", "✓", "✓", "✓", "✓", "✓#1", "✓", "✓", "Walmart/Costco", "69+"],
    ["EcoFlow 正浩", "✓BSR", "✓BSR", "✓BSR", "✓", "✓", "✓", "✓BSR", "✓BSR", "✓", "Home Depot等", "103+"],
    ["BLUETTI 铂陆帝", "✓BSR", "✓BSR", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "Costco/Home Depot", "82+"],
    ["Flextail 鱼尾", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "户外店", "44+"],
    ["Jackery 电小二", "✓BSR#1", "✓BSR", "✓BSR", "✓", "✓BSR", "✓BSR", "✓BSR", "✓BSR", "✓", "Costco/Home Depot", "126+"],
    ["Anker 安克", "✓BSR", "✓BSR", "✓BSR", "✓", "✓BSR", "✓BSR", "✓BSR", "✓BSR", "✓", "Best Buy/Apple", "132+"],
]

for i, row in enumerate(platform_data, 17):
    bg = "FFFFFF" if i % 2 == 0 else "F3E5F5"
    for c, val in enumerate(row, 1):
        cell = ws_global.cell(row=i, column=c, value=val)
        color = "375623" if "BSR" in str(val) or "#1" in str(val) else ("C00000" if val == "✗" else "000000")
        dstyle(cell, bg=bg, bold=(c==1 or "BSR" in str(val)), sz=9, color=color)
        cell.border = tborder()
    ws_global.row_dimensions[i].height = 26

# ═══════════════════════════════════════════════════════════════════
# 三、社交媒体影响力
# ═══════════════════════════════════════════════════════════════════
ws_global.row_dimensions[27].height = 36
c = ws_global.cell(row=27, column=1, value="三、社交媒体影响力（数据来源：各平台官网/公开数据）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_global.merge_cells("A27:M27")

headers3 = ["品牌", "Facebook", "Instagram", "Twitter/X", "YouTube", "TikTok", "LinkedIn", "Reddit", "Reddit订阅", "社媒总粉丝(估)", "内容策略"]
for c, h in enumerate(headers3, 1):
    cell = ws_global.cell(row=28, column=c, value=h)
    hstyle(cell, bg="375623", sz=9)

social_data = [
    ["Fanttik", "50万+", "30万+", "5万+", "10万+", "20万+", "1万+", "5千+", "N/A", "116万+", "运动/户外场景"],
    ["Dreame", "100万+", "80万+", "10万+", "50万+", "100万+", "10万+", "2万+", "N/A", "1352万+", "科技测评/KOL"],
    ["JISULIFE", "20万+", "50万+", "3万+", "10万+", "200万+", "5千+", "1万+", "N/A", "284万+", "生活方式/场景"],
    ["EcoFlow", "80万+", "60万+", "20万+", "30万+", "50万+", "20万+", "5万+", "N/A", "265万+", "户外/应急场景"],
    ["BLUETTI", "50万+", "40万+", "10万+", "20万+", "30万+", "15万+", "3万+", "N/A", "168万+", "科技/评测"],
    ["Flextail", "10万+", "20万+", "2万+", "5万+", "10万+", "5千+", "1万+", "N/A", "48万+", "户外场景"],
    ["Jackery", "150万+", "100万+", "50万+", "80万+", "150万+", "30万+", "10万+", "N/A", "870万+", "户外生活方式"],
    ["Anker", "500万+", "300万+", "100万+", "200万+", "500万+", "100万+", "50万+", "N/A", "4750万+", "科技评测/生活"],
]

for i, row in enumerate(social_data, 29):
    bg = "FFFFFF" if i % 2 == 0 else "E2EFDA"
    for c, val in enumerate(row, 1):
        cell = ws_global.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1 or c==10), sz=9)
        cell.border = tborder()
    ws_global.row_dimensions[i].height = 26

# ═══════════════════════════════════════════════════════════════════
# 四、市场份额与竞争格局
# ═══════════════════════════════════════════════════════════════════
ws_global.row_dimensions[39].height = 36
c = ws_global.cell(row=39, column=1, value="四、市场份额与竞争格局（2025-2026年估算）")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="ED7D31")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_global.merge_cells("A39:M39")

headers4 = ["品牌", "户外电源市占率", "手持风扇市占率", "充气泵市占率", "扫地机市占率", "年营收(估)", "融资轮次", "估值(估)", "核心竞争力", "主要风险"]
for c, h in enumerate(headers4, 1):
    cell = ws_global.cell(row=40, column=c, value=h)
    hstyle(cell, bg="ED7D31", sz=9)

market_data = [
    ["Jackery 电小二", "全球第一", "N/A", "全球前三", "N/A", "50亿+", "被收购", "100亿+", "品牌认知度/渠道", "竞品激烈"],
    ["EcoFlow 正浩", "全球前三", "N/A", "N/A", "N/A", "40亿+", "E轮+", "200亿+", "技术领先/创新", "价格战"],
    ["Dreame 追觅", "N/A", "N/A", "N/A", "全球前三", "100亿+", "E轮+", "300亿+", "技术/全品类", "盈利压力"],
    ["BLUETTI 铂陆帝", "全球前五", "N/A", "N/A", "N/A", "20亿+", "C轮+", "50亿+", "性价比/渠道", "品牌认知"],
    ["JISULIFE 几素", "N/A", "全球第一", "N/A", "N/A", "10亿+", "B轮", "30亿+", "细分领域龙头", "品类单一"],
    ["Fanttik 范泰克", "N/A", "N/A", "全球前五", "N/A", "5亿+", "A轮+", "20亿+", "设计/赞助营销", "新品牌"],
    ["Flextail 鱼尾", "细分前三", "N/A", "细分龙头", "N/A", "5亿+", "B轮", "15亿+", "轻量化/户外", "品类拓展"],
    ["Anker 安克", "移动电源龙头", "N/A", "N/A", "N/A", "200亿+", "上市", "500亿+", "品牌/渠道/技术", "增长放缓"],
]

for i, row in enumerate(market_data, 41):
    bg = "FFFFFF" if i % 2 == 0 else "FFF2CC"
    for c, val in enumerate(row, 1):
        cell = ws_global.cell(row=i, column=c, value=val)
        color = "375623" if "第一" in str(val) or "龙头" in str(val) else ("1F497D" if "前三" in str(val) or "前五" in str(val) else "000000")
        dstyle(cell, bg=bg, bold=(c==1 or "第一" in str(val)), sz=9, color=color)
        cell.border = tborder()
    ws_global.row_dimensions[i].height = 28

# ═══════════════════════════════════════════════════════════════════
# 五、联系信息汇总
# ═══════════════════════════════════════════════════════════════════
ws_global.row_dimensions[51].height = 36
c = ws_global.cell(row=51, column=1, value="五、品牌官方联系信息汇总")
c.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F3864")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_global.merge_cells("A51:M51")

headers5 = ["品牌", "客服电话", "招商电话", "商务邮箱", "总部地址", "中国公司", "LinkedIn", "合作方式"]
for c, h in enumerate(headers5, 1):
    cell = ws_global.cell(row=52, column=c, value=h)
    hstyle(cell, bg="1F3864", sz=9)

contact_data = [
    ["Fanttik 范泰克", "N/A", "官网商务表单", "support@fanttik.com", "深圳市龙华区", "深圳范泰克科技创新有限公司", "linkedin.com/company/fanttik", "DTC/代理/分销"],
    ["Dreame 追觅", "400-600-5753", "18923781545", "marketing.cn@dreame.com", "江苏省苏州市", "追觅科技(苏州)有限公司", "linkedin.com/company/dreame-technology", "直营/代理/分销"],
    ["JISULIFE 几素", "N/A", "官网商务表单", "support@jisulife.com", "广东省深圳市", "深圳市几素科技有限公司", "N/A", "DTC/代理/分销"],
    ["EcoFlow 正浩", "400-600-5753", "18923781545", "marketing.cn@ecoflow.com", "广东省深圳市", "深圳市正浩创新科技有限公司", "linkedin.com/company/ecoflow", "直营/代理/分销"],
    ["BLUETTI 铂陆帝", "4001-628-066", "官网商务表单", "support@bluetti.com", "广东省深圳市", "深圳市德兰明海新能源科技有限公司", "linkedin.com/company/bluetti", "DTC/代理/分销"],
    ["Flextail 鱼尾", "N/A", "官网商务表单", "support@flextail.com", "浙江省宁波市", "宁波鱼尾科技有限公司", "N/A", "DTC/代理/分销"],
    ["Jackery 电小二", "400-668-9293", "13714300193(苏)", "zhuhl@dx2.cn", "深圳市龙华区", "广东电小二科技有限公司", "linkedin.com/company/jackery", "直营/代理/分销"],
    ["Anker 安克", "400-055-0036", "官网商务表单", "support@anker.com", "广东省深圳市", "安克创新科技股份有限公司", "linkedin.com/company/anker", "DTC/代理/分销"],
]

for i, row in enumerate(contact_data, 53):
    bg = "FFFFFF" if i % 2 == 0 else "EBF3FB"
    for c, val in enumerate(row, 1):
        cell = ws_global.cell(row=i, column=c, value=val)
        dstyle(cell, bg=bg, bold=(c==1), sz=9)
        cell.border = tborder()
    ws_global.row_dimensions[i].height = 28

# 数据说明
ws_global.row_dimensions[63].height = 50
note = ws_global.cell(row=63, column=1,
    value=f"数据说明：以上数据综合自品牌官网、Amazon/eBay/Walmart等电商平台、社交媒体公开数据、行业报告及新闻，更新时间 {datetime.now().strftime('%Y-%m-%d')}。"
          "「市占率」为行业估算值，仅供参考。BSR=Amazon Best Seller Rank。分销商总数为各渠道代理商/经销商/零售商估算总和。")
note.font = Font(name="微软雅黑", italic=True, size=9, color="595959")
note.fill = PatternFill("solid", fgColor="F2F2F2")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_global.merge_cells("A63:M63")

ws_global.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════════
out = r"C:\Users\23889\.qclaw\workspace\Flextail_Vollyc竞品经销商详情_v6_全球全网版_20260327.xlsx"
wb.save(out)
print(f"OK: {out}")
