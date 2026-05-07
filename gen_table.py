from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

ws1 = wb.active
ws1.title = '广交会客户背调总表'

hf = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
hfill = PatternFill('solid', fgColor='2F5496')
s_fill = PatternFill('solid', fgColor='FFF2CC')
a_fill = PatternFill('solid', fgColor='DAEEF3')
b_fill = PatternFill('solid', fgColor='E2EFDA')
nf = Font(name='Microsoft YaHei', size=10)
wrap = Alignment(wrap_text=True, vertical='center')
ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

headers = ['序号','客户姓名','公司/品牌','地区','客户级别','声称身份','感兴趣产品','身份验证','背调发现','跟进建议','优先跟进']
for c, h in enumerate(headers, 1):
    cell = ws1.cell(1, c, h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = ctr
    cell.border = bd

data = [
    [1,'Ouou (欧爱梅)','—','北美','S','沃尔玛渠道采购','EVO SUP PUMP 200, TINY TIRE PUMP 600, MAX SHOWER, ZERO POWER 10000C','未验证','未搜到与沃尔玛采购相关公开信息；沃尔玛采购有严格体系（采办会6-10人小组制）','索要@walmart.com邮箱验证；确认是正式员工还是中间商','第一梯队'],
    [2,'丁丁','—','日本','S','大型商超渠道 500+门店','ZERO PUMP 2, TINY PUMP 3X, HELIO 600Z, TINY REPELLER S','未验证','名称过于通用，未找到对应商超；日本500+门店规模商超不多','确认具体商超主体（Don Quijote/AEON？）；核实门店规模','第一梯队'],
    [3,'建伟','—','巴西','S','跨境电商矩阵','TINY PUMP 2X, MAX PUMP 3, ZERO POWER BANK 5000C','未验证','巴西最大电商市场（Mercado Libre为主），名称通用无法定位','确认运营平台和年GMV；索要店铺链接验证实力','第一梯队'],
    [4,'Y.H (友望超市)','—','日本','A','连锁超市总部','TINY REPELLER S, HELIO 600Z, ZERO LANTERN, TINY SHOWER','未验证','未搜到"友望超市"相关公开信息，可能是区域型连锁','确认门店数量和覆盖区域','第二梯队'],
    [5,'White Bai','—','北美','A','渠道分销 电源/充气泵','ZERO POWER 10000C, EVO SUP PUMP PRO, TINY TIRE PUMP 600','未验证','未找到具体公司信息','确认分销渠道类型（线上/线下）和覆盖范围','第二梯队'],
    [6,'楊世豪','青鳥家居 Bluebird Bedding','台湾','A','专业家居连锁','VILLA LANTERN 2, ZERO LANTERN, TINY PILLOW, TINY REPELLER S','已验证','官网bluebirdbedding.com.tw确认存在；台湾寝具品牌，有线上业务','优先跟进，信誉好；台湾市场匹配TINY/ZERO系列','第一梯队'],
    [7,'Peter','sleepEZ','澳洲','A','户外/睡眠品牌','TINY SLEEPING PAD R09 AVS, ZERO PILLOW, ZERO PUMP 2','未验证','未找到sleepEZ品牌明确信息','确认是否注册品牌和销售渠道；产品匹配度极高','第二梯队'],
    [8,'Michael Liss','—','美国CA','A','北美客户','TINY PUMP 3X, ZERO PUMP 2, MAX PUMP 2 PLUS, EVO SUP PUMP 200','未验证','未找到与户外电器分销相关信息','确认公司名称和业务范围；全系列需求说明较专业','第二梯队'],
    [9,'Ali','—','南非','A','零售渠道 5家店铺','TINY PUMP 2X, MAX PUMP 2 PLUS, HELIO 600Z','未验证','名称通用，5家店铺规模较小','规模有限但品类匹配，可给试用订单作南非切入点','第三梯队'],
    [10,'Brandon Hitt','—','美国','B','户外电器','全系列','未验证','信息不足','确认是电商卖家还是分销商','第三梯队'],
    [11,'Adela','MagicLab Robotics','—','B','科技公司','AI功能产品','未验证','信息不足','科技跨界合作有创新潜力，确认采购规模','第三梯队'],
    [12,'Frank','—','—','B','品牌合作 co-brand','3C数码联名','未验证','信息不足','评估品牌溢价能力','第三梯队'],
    [13,'IA international','—','沙特KSA','B','国际公司','耐高温产品','未验证','未找到具体公司信息','中东市场有价值，确认公司实体和采购规模','第三梯队'],
]

for r, row in enumerate(data, 2):
    level = row[4]
    fill = s_fill if level == 'S' else (a_fill if level == 'A' else b_fill)
    for c, val in enumerate(row, 1):
        cell = ws1.cell(r, c, val)
        cell.font = nf
        cell.alignment = wrap if c >= 7 else ctr
        cell.fill = fill
        cell.border = bd

col_widths = [6, 18, 24, 10, 10, 22, 40, 10, 40, 42, 12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Sheet 2: 总体评估
ws2 = wb.create_sheet('总体评估')
ws2.cell(1,1,'广交会客户背调总体评估').font = Font(name='Microsoft YaHei', bold=True, size=14)
ws2.merge_cells('A1:D1')

for c, h in enumerate(['级别','已验证可靠','部分可信','信息不足需核实'], 1):
    cell = ws2.cell(3, c, h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = ctr
    cell.border = bd

for r, row in enumerate([['S级 (3人)',0,0,3],['A级 (6人)',1,0,5],['B级 (5人)',0,0,5],['合计 (14人)',1,0,13]], 4):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(r, c, val)
        cell.font = nf
        cell.alignment = ctr
        cell.border = bd

ws2.cell(9,1,'核心发现').font = Font(name='Microsoft YaHei', bold=True, size=11)
findings = [
    '1. 大多数客户身份无法通过网络验证，广交会客户以微信名为主',
    '2. 唯一已验证公司：青鳥家居（台湾），有官网和真实业务',
    '3. S级客户声称身份需重点核实（沃尔玛采购、500+门店等）',
    '4. 日本市场客户最多，建议专门制定日本市场跟进策略',
    '5. 产品匹配度最高方向：睡眠系统组合、便携充气系列',
]
for i, f in enumerate(findings):
    ws2.cell(10+i, 1, f).font = nf
for i in range(1, 5):
    ws2.column_dimensions[get_column_letter(i)].width = 20

# Sheet 3: 跟进优先级
ws3 = wb.create_sheet('跟进优先级')
ws3.cell(1,1,'客户跟进优先级排序').font = Font(name='Microsoft YaHei', bold=True, size=14)
ws3.merge_cells('A1:C1')

for c, h in enumerate(['优先级','客户','跟进要点'], 1):
    cell = ws3.cell(3, c, h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = ctr
    cell.border = bd

priorities = [
    ['第一梯队','楊世豪（青鳥家居）','已验证可靠，台湾市场，优先报价和样品寄送'],
    ['第一梯队','Ouou（欧爱梅）','声称沃尔玛采购，索要@walmart.com邮箱验证身份'],
    ['第一梯队','丁丁','声称日本500+门店商超，确认具体商超主体'],
    ['第一梯队','建伟','巴西跨境电商，索要店铺链接验证GMV'],
    ['第二梯队','White Bai','北美渠道分销，确认渠道类型和覆盖范围'],
    ['第二梯队','Peter (sleepEZ)','澳洲睡眠品牌，产品匹配度极高'],
    ['第二梯队','Michael Liss','全系列需求，确认采购性质'],
    ['第二梯队','Y.H (友望超市)','日本连锁超市，确认门店规模'],
    ['第三梯队','Ali','南非5店，小规模但品类匹配，可试用订单'],
    ['第三梯队','Brandon Hitt','信息不足，需补充'],
    ['第三梯队','Adela / Frank / IA','信息不足，需进一步沟通'],
]
for r, row in enumerate(priorities, 4):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(r, c, val)
        cell.font = nf
        cell.alignment = wrap
        cell.border = bd

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 55

output_path = r'C:\Users\23889\.qclaw\workspace\广交会客户背调总表_20260428.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
